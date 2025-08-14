from pathlib import Path
import sqlite3
import re
from collections import Counter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / 'Source_Files'
DB = SRC / 'processed_consolidado.db'
RULES_XLSX = SRC / 'Text_Processing_Rules.xlsx'

# This analyzer scans the normalized_descripcion column and finds tokens
# near direction/location adjectives, reporting nouns not present in Noun_Gender.


def _strip_accents(s: str) -> str:
    import unicodedata
    s = unicodedata.normalize('NFKD', s)
    return ''.join(ch for ch in s if not unicodedata.combining(ch))


def load_noun_gender_from_xlsx(path: Path) -> set[str]:
    import openpyxl
    if not path.exists():
        return set()
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Noun_Gender' not in wb.sheetnames:
        return set()
    sh = wb['Noun_Gender']
    nouns = set()
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        nouns.add(_strip_accents(str(row[0]).strip().lower()))
    return nouns


def main(limit: int = 200000):
    if not DB.exists():
        print(f"DB not found: {DB}")
        return
    known = load_noun_gender_from_xlsx(RULES_XLSX)
    con = sqlite3.connect(DB)
    cur = con.cursor()
    cur.execute("PRAGMA journal_mode=OFF")
    cur.execute("PRAGMA temp_store=MEMORY")

    # Pull normalized_descripcion
    q = f"SELECT normalized_descripcion FROM processed_consolidado WHERE normalized_descripcion IS NOT NULL LIMIT {limit}"
    DIR = set(['izquierdo','izquierda','derecho','derecha','delantero','delantera','trasero','trasera'])
    right_after = Counter()
    left_before = Counter()
    for (t,) in cur.execute(q):
        s = (t or '').lower().replace('.', ' ').replace('/', ' ')
        toks = re.findall(r"[a-záéíóúñ0-9]+", s)
        for i, tok in enumerate(toks):
            if tok in DIR:
                if i+1 < len(toks):
                    right_after[_strip_accents(toks[i+1])] += 1
                if i-1 >= 0:
                    left_before[_strip_accents(toks[i-1])] += 1

    # Prefer the noun on the right, but report both, minus known list
    def top_unknown(cnt: Counter, topn=50):
        out = []
        for w, c in cnt.most_common(topn*3):
            if w in known:
                continue
            # Skip obvious adjectives/directions
            if w in DIR or w.endswith('o') or w.endswith('a'):
                pass  # still show, user can decide
            out.append((w, c))
            if len(out) >= topn:
                break
        return out

    print("Top candidates (right after adjective, likely head noun):")
    for w, c in top_unknown(right_after, 80):
        print(f"{w},{c}")

    print("\nTop candidates (left before adjective):")
    for w, c in top_unknown(left_before, 80):
        print(f"{w},{c}")


if __name__ == '__main__':
    main()

