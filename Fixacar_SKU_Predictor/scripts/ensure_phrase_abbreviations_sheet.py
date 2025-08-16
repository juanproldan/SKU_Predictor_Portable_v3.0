#!/usr/bin/env python3
"""
Ensure Text_Processing_Rules.xlsx contains a new sheet 'Abbreviations_Phrases'
for phrase-level expansions. Does not overwrite existing entries.

Sheet schema (header row):
  From | To

Example rows:
  tra d   | trasero derecho
  dl d    | delantero derecho

Run:
  python Fixacar_SKU_Predictor/scripts/ensure_phrase_abbreviations_sheet.py
"""
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / 'Source_Files' / 'Text_Processing_Rules.xlsx'


def main():
    if not XLSX.exists():
        raise SystemExit(f"Rules file not found: {XLSX}")
    wb = openpyxl.load_workbook(XLSX)
    if 'Abbreviations_Phrases' not in wb.sheetnames:
        sh = wb.create_sheet('Abbreviations_Phrases')
        sh.cell(1, 1, 'From')
        sh.cell(1, 2, 'To')
        # seed with a couple of common patterns (can be removed/edited by experts)
        sh.cell(2, 1, 'tra d')
        sh.cell(2, 2, 'trasero derecho')
        sh.cell(3, 1, 'dl d')
        sh.cell(3, 2, 'delantero derecho')
        sh.cell(4, 1, 'del d')
        sh.cell(4, 2, 'delantero derecho')
        print("Created sheet 'Abbreviations_Phrases' with initial examples.")
    else:
        print("Sheet 'Abbreviations_Phrases' already present. No changes.")
    wb.save(XLSX)


if __name__ == '__main__':
    main()

