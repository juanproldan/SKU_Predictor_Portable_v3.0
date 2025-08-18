#!/usr/bin/env python3
"""
Promote vetted (Word, Abbr. 1) pairs from NewAbbreviations to Abbreviations.
Also harden the workflow:
- Idempotent promotions
- Collision detection for phrase pairs (Abbreviations_Phrases: From_Phrase has multiple To_Phrase) → log and skip
- Append Promotions_Log sheet with timestamp/from→to

Run:
  python Fixacar_SKU_Predictor/scripts/promote_abbreviations_pairs.py
"""
from __future__ import annotations
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Set, Tuple
import os

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / 'Source_Files'
XLSX = SRC / 'Text_Processing_Rules.xlsx'
LOGS = ROOT / 'logs'
LOGS.mkdir(exist_ok=True)
PHRASE_LOG = LOGS / '02_new_abbreviations_pairs.log'


def ensure_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def collect_existing(sh_abbr) -> Dict[str, Set[str]]:
    existing: Dict[str, Set[str]] = {}
    for r in range(2, sh_abbr.max_row + 1):
        word = (sh_abbr.cell(r, 1).value or '').strip().lower()
        if not word:
            continue
        c = 2
        abbrs: Set[str] = set()
        while c <= sh_abbr.max_column and (sh_abbr.cell(1, c).value or '').strip().lower().startswith('abbr'):
            v = (sh_abbr.cell(r, c).value or '').strip().lower()
            if v:
                abbrs.add(v)
            c += 1
        if word:
            existing[word] = existing.get(word, set()) | abbrs
    return existing


def append_abbr(sh_abbr, word: str, abbr: str):
    # Find row for word or create
    row_idx = None
    for r in range(2, sh_abbr.max_row + 1):
        if (sh_abbr.cell(r, 1).value or '').strip().lower() == word:
            row_idx = r
            break
    if row_idx is None:
        row_idx = sh_abbr.max_row + 1
        sh_abbr.cell(row_idx, 1, word)
        # Ensure header exists
        if (sh_abbr.cell(1, 2).value or '').strip() == '':
            sh_abbr.cell(1, 2, 'Abbr. 1')
    # Skip if already present in row (idempotent)
    c = 2
    while c <= sh_abbr.max_column and (sh_abbr.cell(1, c).value or '').strip().lower().startswith('abbr'):
        if (sh_abbr.cell(row_idx, c).value or '').strip().lower() == abbr:
            return
        c += 1
    # Find next empty Abbr.* column
    c = 2
    while c <= sh_abbr.max_column and (sh_abbr.cell(row_idx, c).value or ''):
        c += 1
    sh_abbr.cell(row_idx, c, abbr)



def detect_phrase_collisions(wb) -> List[tuple[str, list[str]]]:
    if 'Abbreviations_Phrases' not in wb.sheetnames:
        return []
    sh = wb['Abbreviations_Phrases']
    headers = { (str(sh.cell(1, c).value or '').strip().lower()): c for c in range(1, sh.max_column+1) }
    cf = headers.get('from_phrase', 1)
    ct = headers.get('to_phrase', 2)
    cc = headers.get('confidence', None)
    allowed = {'high', 'medium'}
    m: Dict[str, Set[str]] = {}
    for r in range(2, sh.max_row + 1):
        conf = (str(sh.cell(r, cc).value or '').strip().lower()) if cc else ''
        if cc and conf not in allowed:
            continue
        f = (sh.cell(r, cf).value or '').strip().lower()
        t = (sh.cell(r, ct).value or '').strip().lower()
        if f and t:
            m.setdefault(f, set()).add(t)
    return [(f, sorted(list(to_set))) for f, to_set in m.items() if len(to_set) > 1]


def log_collisions(collisions: List[tuple[str, list[str]]], log_path: Path):
    if not collisions:
        return
    with open(log_path, 'a', encoding='utf-8') as f:
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for fph, to_list in collisions:
            f.write(f"[{ts}] COLLISION From_Phrase='{fph}' -> {to_list}\n")


def promote():
    if not XLSX.exists():
        raise SystemExit(f"Rules file not found: {XLSX}")

    wb = openpyxl.load_workbook(XLSX)
    sh_new = ensure_sheet(wb, 'NewAbbreviations')
    sh_hist = ensure_sheet(wb, 'NewAbbreviations_History')
    sh_abbr = ensure_sheet(wb, 'Abbreviations')

    # Ensure history and rejects headers
    if sh_hist.max_row == 1:
        for c in range(1, sh_new.max_column + 1):
            sh_hist.cell(1, c, sh_new.cell(1, c).value)
        sh_hist.cell(1, sh_new.max_column + 1, 'Promoted_At')
    sh_rej = ensure_sheet(wb, 'NewAbbreviations_Rejects')
    if sh_rej.max_row == 1:
        sh_rej.cell(1, 1, 'Word')
        sh_rej.cell(1, 2, 'Abbr. 1')
        sh_rej.cell(1, 3, 'Rejected_At')

    # Existing pairs
    existing = collect_existing(sh_abbr)

    # Find columns
    headers = { (sh_new.cell(1, c).value or '').strip().lower(): c for c in range(1, sh_new.max_column + 1) }
    col_word = headers.get('word', 1)
    col_abbr = headers.get('abbr. 1', 2)
    col_approve = headers.get('approve?')
    if not col_approve:
        raise SystemExit("NewAbbreviations sheet is missing 'Approve?' column.")

    promoted_rows: List[int] = []
    rejected_rows: List[int] = []

    # Collision detection for phrase pairs
    collisions = detect_phrase_collisions(wb)
    log_collisions(collisions, PHRASE_LOG)

    for r in range(2, sh_new.max_row + 1):
        approve = (sh_new.cell(r, col_approve).value or '').strip().lower()
        word = (sh_new.cell(r, col_word).value or '').strip().lower()
        abbr = (sh_new.cell(r, col_abbr).value or '').strip().lower()

        if approve in {'y', 'yes'}:
            if not abbr or not word:
                continue
            if abbr in existing.get(word, set()):
                promoted_rows.append(r)
                continue
            append_abbr(sh_abbr, word, abbr)
            existing.setdefault(word, set()).add(abbr)
            promoted_rows.append(r)
        elif approve in {'n', 'no'}:
            # Add to rejects if not already there
            already = False
            for rr in range(2, sh_rej.max_row + 1):
                if (sh_rej.cell(rr, 1).value or '').strip().lower() == word and (sh_rej.cell(rr, 2).value or '').strip().lower() == abbr:
                    already = True
                    break
            if not already:
                dr = sh_rej.max_row + 1
                sh_rej.cell(dr, 1, word)
                sh_rej.cell(dr, 2, abbr)
                sh_rej.cell(dr, 3, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            rejected_rows.append(r)
        else:
            # Skip undecided rows
            continue

    # Append to history with timestamp
    if promoted_rows:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for r in promoted_rows:
            dest_row = sh_hist.max_row + 1
            for c in range(1, sh_new.max_column + 1):
                sh_hist.cell(dest_row, c, sh_new.cell(r, c).value)
            sh_hist.cell(dest_row, sh_new.max_column + 1, timestamp)

    # Promotions_Log sheet
    sh_log = ensure_sheet(wb, 'Promotions_Log')
    if sh_log.max_row == 1:
        sh_log.cell(1,1,'Timestamp'); sh_log.cell(1,2,'Word'); sh_log.cell(1,3,'Abbr')
    for r in promoted_rows:
        rr = sh_log.max_row + 1
        sh_log.cell(rr, 1, timestamp)
        sh_log.cell(rr, 2, (sh_new.cell(r, col_word).value or '').strip().lower())
        sh_log.cell(rr, 3, (sh_new.cell(r, col_abbr).value or '').strip().lower())

    # Delete both promoted and rejected rows from NewAbbreviations
    for r in sorted(set(promoted_rows + rejected_rows), reverse=True):
        sh_new.delete_rows(r, 1)

    wb.save(XLSX)
    print(f"Promotion complete. Promoted rows: {len(promoted_rows)}, Rejected rows: {len(rejected_rows)}")


if __name__ == '__main__':
    promote()

