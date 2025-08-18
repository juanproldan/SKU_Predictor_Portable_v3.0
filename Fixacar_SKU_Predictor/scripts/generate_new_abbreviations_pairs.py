#!/usr/bin/env python3
"""
Generate abbreviation pairs (Word, Abbr. 1) and write into
Text_Processing_Rules.xlsx sheet 'NewAbbreviations'.

Differences vs generate_new_abbreviations.py:
- One row per (word, abbreviation) pair (words may repeat across rows)
- Columns: Word | Abbr. 1 | Normalized_descripcion | frequency | Normalized_descripcion02 | frequency02 | Approve?
- Dedups against existing pairs in Abbreviations and NewAbbreviations
- If the word cannot be confidently inferred, Word is left blank (abbreviation is still proposed)
- Sheet is cleared on each run to regenerate fresh suggestions (no append)
- Respects NewAbbreviations_Rejects for exclusions

Run:
  python Fixacar_SKU_Predictor/scripts/generate_new_abbreviations_pairs.py
"""
from __future__ import annotations
from pathlib import Path
import sqlite3
import re
from collections import Counter, defaultdict
from typing import Dict, List, Set, Tuple

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / 'Source_Files'
DB = SRC / 'processed_consolidado.db'
XLSX = SRC / 'Text_Processing_Rules.xlsx'

TOKEN_RE = re.compile(r"[a-záéíóúñ0-9]+", re.IGNORECASE)
ONLY_LETTERS_RE = re.compile(r"^[a-záéíóúñ]+$")

# Strong domain mappings (prefixes -> canonical word) [no single-letter keys]
DIR_PREFIX_TO_WORD = {
    # izquierda
    'iz': 'izquierda', 'izq': 'izquierda', 'izqu': 'izquierda', 'izqui': 'izquierda',
    # derecha
    'der': 'derecha', 'dere': 'derecha', 'derec': 'derecha',
    # delantero
    'ant': 'delantero', 'del': 'delantero', 'dela': 'delantero', 'delan': 'delantero', 'delant': 'delantero',
    # trasero
    'tra': 'trasero', 'tras': 'trasero', 'post': 'trasero',
    # superior / inferior
    'sup': 'superior', 'inf': 'inferior',
}

STOPWORDS = {
    'de', 'del', 'la', 'el', 'los', 'las', 'y', 'o', 'u', 'con', 'sin', 'para', 'por',
}


def read_all_tokens() -> Tuple[Counter, Dict[str, Counter], Dict[str, Counter], Counter]:
    if not DB.exists():
        raise SystemExit(f"DB not found: {DB}")
    con = sqlite3.connect(DB)
    cur = con.cursor()
    q = "SELECT normalized_descripcion FROM processed_consolidado WHERE normalized_descripcion IS NOT NULL"

    token_counts: Counter = Counter()
    window = 2
    contexts: Dict[str, Counter] = defaultdict(Counter)
    token_descs: Dict[str, Counter] = defaultdict(Counter)
    desc_counts: Counter = Counter()

    for (text,) in cur.execute(q):
        if not text:
            continue
        s_orig = str(text).lower().strip()
        desc_counts[s_orig] += 1
        s = s_orig.replace('.', ' ').replace('/', ' ')
        toks = TOKEN_RE.findall(s)
        toks = [t for t in toks if t]
        token_counts.update(toks)
        # contexts
        n = len(toks)
        for i, t in enumerate(toks):
            start = max(0, i - window)
            end = min(n, i + window + 1)
            for j in range(start, end):
                if j == i:
                    continue
                contexts[t][toks[j]] += 1
        for t in set(toks):
            token_descs[t][s_orig] += 1

    con.close()
    return token_counts, contexts, token_descs, desc_counts


def read_existing_pairs(wb) -> Tuple[Set[Tuple[str, str]], Dict[str, Set[str]]]:
    """Collect existing (word, abbr) pairs from Abbreviations and NewAbbreviations.
    Returns:
      existing_pairs: set of (word_or_blank, abbr)
      abbr_by_word: word -> set(abbr)
    """
    existing_pairs: Set[Tuple[str, str]] = set()
    abbr_by_word: Dict[str, Set[str]] = defaultdict(set)

    # Abbreviations sheet
    if 'Abbreviations' in wb.sheetnames:
        sh = wb['Abbreviations']
        for r in range(2, sh.max_row + 1):
            word = (sh.cell(r, 1).value or '').strip().lower()
            if not word:
                continue
            c = 2
            while c <= sh.max_column and (sh.cell(1, c).value or '').strip().lower().startswith('abbr'):
                abbr = (sh.cell(r, c).value or '').strip().lower()
                if abbr:
                    existing_pairs.add((word, abbr))
                    abbr_by_word[word].add(abbr)
                c += 1

    # NewAbbreviations sheet
    if 'NewAbbreviations' in wb.sheetnames:
        sh = wb['NewAbbreviations']
        for r in range(2, sh.max_row + 1):
            word = (sh.cell(r, 1).value or '').strip().lower()
            abbr = (sh.cell(r, 2).value or '').strip().lower()
            if abbr:
                existing_pairs.add((word, abbr))
                if word:
                    abbr_by_word[word].add(abbr)

    return existing_pairs, abbr_by_word


def load_or_create_new_sheet(wb):
    if 'NewAbbreviations' not in wb.sheetnames:
        sh = wb.create_sheet('NewAbbreviations')
        sh.cell(1, 1, 'Word')
        sh.cell(1, 2, 'Abbr. 1')
        sh.cell(1, 3, 'Normalized_descripcion')
        sh.cell(1, 4, 'frequency')
        sh.cell(1, 5, 'Normalized_descripcion02')
        sh.cell(1, 6, 'frequency02')
        sh.cell(1, 7, 'Approve?')
    else:
        sh = wb['NewAbbreviations']
        # Ensure headers exist
        headers = { (sh.cell(1, c).value or '').strip().lower(): c for c in range(1, sh.max_column + 1) }
        if 'word' not in headers: sh.cell(1, 1, 'Word')
        if 'abbr. 1' not in headers: sh.cell(1, 2, 'Abbr. 1')
        if 'normalized_descripcion' not in headers: sh.cell(1, 3, 'Normalized_descripcion')
        if 'frequency' not in headers: sh.cell(1, 4, 'frequency')
        if 'normalized_descripcion02' not in headers: sh.cell(1, 5, 'Normalized_descripcion02')
        if 'frequency02' not in headers: sh.cell(1, 6, 'frequency02')
        if 'approve?' not in headers: sh.cell(1, 7, 'Approve?')
    return sh


def build_canonical_pool(wb, vocab_counts: Counter) -> Set[str]:
    pool: Set[str] = set()
    if 'Equivalencias' in wb.sheetnames:
        for row in wb['Equivalencias'].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                t = str(row[0]).strip().lower()
                if ONLY_LETTERS_RE.match(t):
                    pool.add(t)
    if 'Noun_Gender' in wb.sheetnames:
        for row in wb['Noun_Gender'].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                t = str(row[0]).strip().lower()
                if ONLY_LETTERS_RE.match(t):
                    pool.add(t)
    for w, cnt in vocab_counts.items():
        if len(w) >= 5 and ONLY_LETTERS_RE.match(w) and cnt >= 5:
            pool.add(w)
    return pool


def is_potential_abbr(tok: str, freq: int) -> bool:
    t = tok.lower()
    L = len(t)
    if L < 2 or L > 6:
        return False
    if t in STOPWORDS and t not in {'del'}:
        return False
    if t.isdigit() or re.search(r"[a-z][0-9]|[0-9][a-z]", t):
        return False
    if not ONLY_LETTERS_RE.match(t):
        return False
    return freq >= 3


def pick_word_for_abbr(abbr: str, vocab_counts: Counter, canonical_pool: Set[str], contexts: Dict[str, Counter]) -> Tuple[str | None, float]:
    a = abbr.lower()
    for pfx, target in DIR_PREFIX_TO_WORD.items():
        if a.startswith(pfx) and len(a) <= max(2, len(pfx) + 1):
            return target, 1.0
    cands = []
    for w, cnt in vocab_counts.items():
        if len(w) >= 5 and w.startswith(a) and ONLY_LETTERS_RE.match(w):
            cands.append((w, cnt))
    for w in canonical_pool:
        if w.startswith(a):
            cands.append((w, vocab_counts.get(w, 1)))
    if not cands:
        return None, 0.0
    a_ctx = contexts.get(a)
    def score(w: str, cnt: int) -> float:
        s = float(cnt)
        if a_ctx:
            overlap = 0
            w_ctx = contexts.get(w)
            if w_ctx:
                for tok, c in a_ctx.most_common(10):
                    overlap += min(c, w_ctx.get(tok, 0))
            s += 0.1 * overlap
        return s
    best = max(cands, key=lambda x: score(x[0], x[1]))
    return best[0], float(best[1])


def main():
    if not XLSX.exists():
        raise SystemExit(f"Rules file not found: {XLSX}")

    token_counts, contexts, token_descs, desc_counts = read_all_tokens()

    wb = openpyxl.load_workbook(XLSX)
    sh = load_or_create_new_sheet(wb)

    # Clear existing suggestion rows (keep header)
    if sh.max_row > 1:
        sh.delete_rows(2, sh.max_row - 1)

    existing_pairs, _ = read_existing_pairs(wb)

    # Collect rejects to exclude
    rejects_set: Set[Tuple[str, str]] = set()
    if 'NewAbbreviations_Rejects' in wb.sheetnames:
        sh_rej = wb['NewAbbreviations_Rejects']
        for r in range(2, sh_rej.max_row + 1):
            w = (sh_rej.cell(r, 1).value or '').strip().lower()
            a = (sh_rej.cell(r, 2).value or '').strip().lower()
            if a:
                rejects_set.add((w, a))

    canonical_pool = build_canonical_pool(wb, token_counts)

    candidates = [t for t, f in token_counts.items() if is_potential_abbr(t, f)]

    # Collect new rows as a list of tuples to allow frequency sorting
    rows: List[Tuple[str, str, str, int, str, int]] = []  # (word, abbr, desc1, freq, desc2, freq2)

    for a in candidates:
        # compute top 2 normalized_descripcion candidates
        desc_counter = token_descs.get(a, Counter())
        top2 = desc_counter.most_common(2)
        best_desc = top2[0][0] if len(top2) >= 1 else ''
        second_desc = top2[1][0] if len(top2) >= 2 else ''
        best_freq = desc_counts.get(best_desc, 0)
        second_freq = desc_counts.get(second_desc, 0)
        # map to word
        word, _ = pick_word_for_abbr(a, token_counts, canonical_pool, contexts)
        if word is not None and word == a:
            word = None
        pair_key = ((word or ''), a)
        if pair_key in existing_pairs:
            continue
        if (word or '', a) in rejects_set:
            continue
        # Add row
        rows.append((word or '', a, best_desc or '', best_freq, second_desc or '', second_freq))

    # Sort rows by frequency desc
    rows.sort(key=lambda r: (r[3], r[2]), reverse=True)

    # Append rows
    for word, abbr, desc1, freq, desc2, freq2 in rows:
        r = sh.max_row + 1
        sh.cell(r, 1, word)
        sh.cell(r, 2, abbr)
        sh.cell(r, 3, desc1)
        sh.cell(r, 4, freq)
        sh.cell(r, 5, desc2)
        sh.cell(r, 6, freq2)
        sh.cell(r, 7, '')  # Approve?

    wb.save(XLSX)
    print(f"NewAbbreviations regenerated in {XLSX}. Rows: {len(rows)}")


if __name__ == '__main__':
    main()

