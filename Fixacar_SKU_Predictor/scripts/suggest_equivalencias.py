#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Suggest candidate synonym/equivalencia pairs from processed_consolidado.db and export
suggestions to a new sheet 'ew_Equivalencias' in Text_Processing_Rules.xlsx.

Notes
- This script DOES NOT modify the Equivalencias sheet used by the app. It only
  creates/overwrites a helper sheet for experts to review.
- It analyzes normalized_descripcion and referencia. Words that map to very similar
  referencia sets (high Jaccard) are proposed as equivalents. We also surface likely
  orthographic variants using difflib similarity.
- No spaCy and no heavy deps. Only stdlib + openpyxl.

Run
  py -3.11 scripts\suggest_equivalencias.py

Optional flags
  --db <path>          Path to processed_consolidado.db (default: Source_Files/...)
  --rules <path>       Path to Text_Processing_Rules.xlsx (default: Source_Files/...)
  --min-freq N         Minimum word frequency to consider (default: 30)
  --min-shared N       Minimum shared referencias between two words (default: 25)
  --jaccard T          Jaccard similarity threshold (default: 0.6)
  --max-words N        Consider top N words by freq (default: 2000)

Output sheet columns
  group_id, canonical, synonym, method, similarity_score,
  word_freq, synonym_freq, shared_ref_count, total_ref_word, total_ref_synonym
"""
from __future__ import annotations

import argparse
import collections
import itertools
import math
import os
import re
import sqlite3
from pathlib import Path
from difflib import SequenceMatcher
from typing import Dict, List, Set, Tuple

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception as e:  # pragma: no cover
    openpyxl = None

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "Source_Files"
DB_PATH = SRC / "processed_consolidado.db"
RULES_XLSX = SRC / "Text_Processing_Rules.xlsx"
SHEET_NAME = "ew_Equivalencias"

# Light Spanish stoplist and non-specific terms to reduce noise
STOP = {
    "del", "de", "la", "el", "los", "las", "y", "o", "u", "en", "con", "sin",
    "para", "por", "un", "una", "uno", "al", "a", "se", "su", "sus", "lo",
    # domain-generic
    "lado", "derecha", "derecho", "izquierda", "izquierdo", "delantero", "trasero",
    "superior", "inferior", "delantera", "trasera", "izq", "der", "del", "post",
    # small tokens
    "mm", "cm", "kit", "set", "par", "x", "paragolpes",  # keep 'paragolpes'? prefer to keep; remove if over-filtering
}
TOKEN_RE = re.compile(r"[a-z0-9]+", re.IGNORECASE)


def tokenize(text: str) -> List[str]:
    if not text:
        return []
    return TOKEN_RE.findall(text.lower())


def load_rows(db_path: Path) -> List[Tuple[str, str]]:
    con = sqlite3.connect(str(db_path))
    cur = con.cursor()
    cur.execute(
        """
        SELECT normalized_descripcion, referencia
        FROM processed_consolidado
        WHERE normalized_descripcion IS NOT NULL AND TRIM(normalized_descripcion) <> ''
          AND referencia IS NOT NULL AND TRIM(referencia) <> ''
        """
    )
    rows = cur.fetchall()
    con.close()
    return rows


def build_inverted_index(rows: List[Tuple[str, str]], min_freq: int) -> Tuple[Dict[str, Set[str]], Dict[str, int]]:
    # word -> set of referencias it appears with
    word_to_refs: Dict[str, Set[str]] = collections.defaultdict(set)
    freq: Dict[str, int] = collections.Counter()

    for desc, ref in rows:
        words = [w for w in tokenize(desc) if len(w) >= 3 and w not in STOP]
        if not words:
            continue
        unique_words = set(words)  # avoid counting same word twice per row
        for w in unique_words:
            word_to_refs[w].add(ref)
        for w in words:
            freq[w] += 1

    # prune by min_freq
    pruned = {w: refs for w, refs in word_to_refs.items() if freq[w] >= min_freq}
    pruned_freq = {w: freq[w] for w in pruned}
    return pruned, pruned_freq


def jaccard(a: Set[str], b: Set[str]) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    if inter == 0:
        return 0.0
    union = len(a | b)
    return inter / union


def sequence_sim(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def candidate_pairs(words: List[str], word_to_refs: Dict[str, Set[str]], min_shared: int) -> List[Tuple[str, str, int]]:
    # Generate pairs that share at least min_shared referencias to avoid O(n^2) blowup
    # Build reverse map: ref -> words
    ref_to_words: Dict[str, List[str]] = collections.defaultdict(list)
    for w in words:
        for r in word_to_refs[w]:
            ref_to_words[r].append(w)

    common_counts: Dict[Tuple[str, str], int] = collections.Counter()
    for words_for_ref in ref_to_words.values():
        words_for_ref.sort()
        for i in range(len(words_for_ref)):
            wi = words_for_ref[i]
            for j in range(i + 1, len(words_for_ref)):
                wj = words_for_ref[j]
                common_counts[(wi, wj)] += 1

    pairs = [(a, b, c) for (a, b), c in common_counts.items() if c >= min_shared]
    return pairs


def find_equivalences(word_to_refs: Dict[str, Set[str]], freq: Dict[str, int], *,
                      max_words: int, min_shared: int, jaccard_thr: float) -> List[Tuple[str, str, float, int, int, int]]:
    # Consider top-N frequent words to bound runtime
    top_words = [w for w, _ in sorted(freq.items(), key=lambda kv: (-kv[1], kv[0]))[:max_words]]

    # Pre-filter pairs by shared refs
    pairs = candidate_pairs(top_words, word_to_refs, min_shared)

    results = []
    for a, b, shared in pairs:
        ja = jaccard(word_to_refs[a], word_to_refs[b])
        if ja >= jaccard_thr:
            results.append((a, b, ja, shared, len(word_to_refs[a]), len(word_to_refs[b])))
            continue
        # Also allow very high textual similarity (likely spelling variants)
        sim = sequence_sim(a, b)
        if sim >= 0.92 and shared >= max(10, min_shared // 2):
            # Treat as an orthographic equivalence even if Jaccard is a bit low
            results.append((a, b, sim, shared, len(word_to_refs[a]), len(word_to_refs[b])))
    return results


def groups_from_pairs(pairs: List[Tuple[str, str, float, int, int, int]], freq: Dict[str, int]):
    # Build undirected graph and take connected components
    parent: Dict[str, str] = {}

    def find(x):
        parent.setdefault(x, x)
        if parent[x] != x:
            parent[x] = find(parent[x])
        return parent[x]

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for a, b, *_ in pairs:
        union(a, b)

    comps: Dict[str, List[str]] = collections.defaultdict(list)
    for w in set(list(freq.keys()) + [w for ab in pairs for w in ab[:2]]):
        if w in parent:
            comps[find(w)].append(w)

    # pick canonical as highest frequency in each component
    groups: Dict[str, Tuple[str, List[str]]] = {}
    for root, words in comps.items():
        canonical = max(words, key=lambda w: (freq.get(w, 0), -len(w)))
        groups[root] = (canonical, sorted([w for w in words if w != canonical]))
    return groups


def write_sheet(xlsx_path: Path, rows_out: List[Tuple[int, str, str, str, float, int, int, int, int, int]]):
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Run 0_install_packages.bat.")

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    # Remove default sheet if it's the only one and empty
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and wb["Sheet"].max_row == 1:
        del wb["Sheet"]

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    headers = [
        "group_id", "canonical", "synonym", "method", "similarity_score",
        "word_freq", "synonym_freq", "shared_ref_count", "total_ref_word", "total_ref_synonym",
    ]
    ws.append(headers)

    for row in rows_out:
        ws.append(list(row))

    # Autosize
    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    try:
        wb.save(xlsx_path)
    except PermissionError as e:
        raise SystemExit(f"[ERROR] Could not write to {xlsx_path}. Is the file open in Excel? Please close it and retry.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", type=Path, default=DB_PATH)
    ap.add_argument("--rules", type=Path, default=RULES_XLSX)
    ap.add_argument("--min-freq", type=int, default=30)
    ap.add_argument("--min-shared", type=int, default=25)
    ap.add_argument("--jaccard", type=float, default=0.60)
    ap.add_argument("--max-words", type=int, default=2000)
    args = ap.parse_args()

    print(f"[INFO] DB: {args.db}")
    print(f"[INFO] Rules: {args.rules}")

    rows = load_rows(args.db)
    print(f"[INFO] Loaded {len(rows):,} rows from DB")

    word_to_refs, freq = build_inverted_index(rows, args.min_freq)
    print(f"[INFO] Candidate words (freq >= {args.min_freq}): {len(word_to_refs):,}")

    pairs = find_equivalences(
        word_to_refs,
        freq,
        max_words=args.max_words,
        min_shared=args.min_shared,
        jaccard_thr=args.jaccard,
    )
    print(f"[INFO] Candidate pairs after filters: {len(pairs):,}")

    groups = groups_from_pairs(pairs, freq)
    print(f"[INFO] Groups: {len(groups):,}")

    # Flatten to rows
    out_rows: List[Tuple[int, str, str, str, float, int, int, int, int, int]] = []
    gid = 1
    for _, (canonical, syns) in sorted(groups.items(), key=lambda kv: kv[0]):
        for a, b, score, shared, ta, tb in pairs:
            if (a == canonical and b in syns) or (b == canonical and a in syns):
                method = "jaccard" if score <= 1.0 else "edit_sim"  # both produce <=1.0; keep label trivial
                # We don't retain which metric created each edge; approximate label
                out_rows.append(
                    (
                        gid,
                        canonical,
                        b if a == canonical else a,
                        method,
                        round(float(score), 4),
                        int(freq.get(canonical, 0)),
                        int(freq.get(b if a == canonical else a, 0)),
                        int(shared),
                        int(ta if a == canonical else tb),
                        int(tb if a == canonical else ta),
                    )
                )
        gid += 1

    if not out_rows:
        print("[WARN] No suggestions found with current thresholds. Try lowering --min-freq or --jaccard.")

    write_sheet(args.rules, out_rows)
    print(f"[OK] Wrote {len(out_rows):,} suggestion rows to sheet '{SHEET_NAME}' in {args.rules}")


if __name__ == "__main__":
    main()

