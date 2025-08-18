#!/usr/bin/env python3
"""
Generate "Series_Group_Candidates (2)" sheet based on VIN10 prefix analysis.
- One row per (VIN10, Maker, Model)
- Columns:
  [VIN_Prefix_Type, VIN_Prefix, Maker, Model, Group?, Candidate_Series01, Count01, ..., Candidate_Series06, Count06]
- Candidate series are ordered by distinct-VIN count (most frequent first)
- Group? left blank for expert to mark (preserved on re-run)
"""
from __future__ import annotations
import os
import sqlite3
from collections import defaultdict, Counter
from typing import Dict, Tuple
import openpyxl

ROOT = os.path.dirname(os.path.dirname(__file__))
DB_PATH = os.path.join(ROOT, 'Source_Files', 'processed_consolidado.db')
XLSX = os.path.join(ROOT, 'Source_Files', 'Text_Processing_Rules.xlsx')

SHEET_NAME = 'Series_Group_Candidates'
HEADERS = [
    'VIN_Prefix_Type',
    'VIN_Prefix',
    'Maker',
    'Model',
    'Group?',
    'Candidate_Series01', 'Count01',
    'Candidate_Series02', 'Count02',
    'Candidate_Series03', 'Count03',
    'Candidate_Series04', 'Count04',
    'Candidate_Series05', 'Count05',
    'Candidate_Series06', 'Count06',
]


def _ensure_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    sh = wb.create_sheet(SHEET_NAME)
    for i, h in enumerate(HEADERS, start=1):
        sh.cell(1, i, h)
    return sh


def _read_existing_group_flags(sh) -> Dict[Tuple[str, str, str], str]:
    """Preserve existing Group? flags keyed by (vin10, maker, model_str)."""
    flags: Dict[Tuple[str, str, str], str] = {}
    if sh.max_row <= 1:
        return flags
    for r in range(2, sh.max_row + 1):
        vin10 = (sh.cell(r, 2).value or '').strip().upper()
        maker = (sh.cell(r, 3).value or '').strip().upper()
        model_val = sh.cell(r, 4).value
        model_str = '' if model_val is None else str(model_val).strip()
        grp = sh.cell(r, 5).value
        if vin10 and maker:
            flags[(vin10, maker, model_str)] = grp
    return flags


def _fetch_counts_by_vin10(con: sqlite3.Connection) -> Dict[Tuple[str, str, str], Counter]:
    """Return mapping: (vin10, maker, model_str) -> Counter(series -> count of distinct VINs)."""
    cur = con.cursor()
    counts: Dict[Tuple[str, str, str], Counter] = defaultdict(Counter)
    q = (
        "SELECT SUBSTR(UPPER(vin_number),1,10) AS vin10, UPPER(maker) AS maker, model, series, COUNT(DISTINCT UPPER(vin_number)) AS c "
        "FROM processed_consolidado "
        "WHERE vin_number IS NOT NULL AND LENGTH(vin_number)=17 AND maker IS NOT NULL AND series IS NOT NULL AND TRIM(series)<>'' "
        "GROUP BY vin10, maker, model, series"
    )
    cur.execute(q)
    for vin10, maker, model_val, series, c in cur.fetchall():
        model_str = '' if model_val is None else str(model_val)
        counts[(vin10, maker, model_str)][series] += int(c)
    return counts


def main():
    if not os.path.exists(DB_PATH):
        raise SystemExit(f"DB not found: {DB_PATH}")
    if not os.path.exists(XLSX):
        raise SystemExit(f"Rules file not found: {XLSX}")

    con = sqlite3.connect(DB_PATH)
    counts = _fetch_counts_by_vin10(con)
    con.close()

    wb = openpyxl.load_workbook(XLSX)
    sh = _ensure_sheet(wb)

    # Preserve existing Group? flags
    existing_flags = _read_existing_group_flags(sh)

    # Clear existing rows (keep header)
    if sh.max_row > 1:
        sh.delete_rows(2, sh.max_row - 1)

    # Emit one row per (vin10, maker, model)
    total_rows = 0
    for (vin10, maker, model_str), cnt in counts.items():
        ranked = sorted(cnt.items(), key=lambda kv: kv[1], reverse=True)
        group_flag = existing_flags.get((vin10, maker, model_str), '')
        row = ['VIN10', vin10, maker, model_str, group_flag]
        # Add up to 6 series/count pairs
        for i in range(6):
            if i < len(ranked):
                series, c = ranked[i]
                row.extend([series, int(c)])
            else:
                row.extend(['', ''])
        sh.append(row)
        total_rows += 1

    wb.save(XLSX)
    print(f"Wrote {SHEET_NAME}: {total_rows} rows")


if __name__ == '__main__':
    main()

