from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / 'Source_Files' / 'Text_Processing_Rules.xlsx'

SEED = [
    ('paragolpes','m'), ('bumper','m'), ('persiana','f'), ('radiador','m'),
    ('compuerta','f'), ('capo','m'), ('estribo','m'), ('tapa','f'),
    ('luz','f'), ('motor','m'), ('placa','f'), ('panel','m')
]

def ensure_sheet_and_seed():
    if not XLSX.exists():
        raise FileNotFoundError(f'Rules file missing: {XLSX}')
    wb = openpyxl.load_workbook(XLSX)
    if 'Noun_Gender' not in wb.sheetnames:
        sh = wb.create_sheet('Noun_Gender')
        sh.cell(1,1,'noun')
        sh.cell(1,2,'gender')
        existing = set()
    else:
        sh = wb['Noun_Gender']
        # ensure headers
        if (sh.cell(1,1).value or '').strip().lower() != 'noun':
            sh.cell(1,1,'noun')
        if (sh.cell(1,2).value or '').strip().lower() != 'gender':
            sh.cell(1,2,'gender')
        existing = set()
        for r in sh.iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                existing.add(str(r[0]).strip().lower())
    added = 0
    for noun, g in SEED:
        if noun not in existing:
            sh.append([noun, g])
            added += 1
    wb.save(XLSX)
    return added

if __name__ == '__main__':
    added = ensure_sheet_and_seed()
    print(f'Updated Noun_Gender; added {added} seed rows if missing.')

