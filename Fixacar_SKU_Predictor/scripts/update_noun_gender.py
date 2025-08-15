#!/usr/bin/env python3
"""
update_noun_gender.py

One-shot script to discover real nouns from normalized_descripcion and append
(noun, gender) to Source_Files/Text_Processing_Rules.xlsx → Noun_Gender.

Design goals:
- Only append known/real nouns (no abbreviations like izq/der/tra/rh/lh, no 'del', 'post', etc.).
- Assign gender using domain knowledge first, then conservative morphology.
- Append only new nouns (do not duplicate nor overwrite), idempotent across runs.
- No external NLP; uses existing project utilities and the rules workbook.

Usage:
  python Fixacar_SKU_Predictor/scripts/update_noun_gender.py --top 8000
"""
from __future__ import annotations
from pathlib import Path
import argparse
import sqlite3
import re
import sys
from collections import Counter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / 'Source_Files'
DB = SRC / 'processed_consolidado.db'
XLSX = SRC / 'Text_Processing_Rules.xlsx'

# Reuse utilities from portable_app
sys.path.append(str((ROOT / 'portable_app' / 'src').resolve()))
from utils.text_utils import get_noun_gender as _gender_from_lists  # 'masculine'/'feminine'

# Tokenization (keep accents for readability during checks)
TOKEN_RE = re.compile(r"[a-z0-9áéíóúñ]+", re.IGNORECASE)

# Spanish direction/location adjectives (normalized forms)
DIR_ADJ = {
    'izquierda', 'derecha', 'delantera', 'trasera', 'superior', 'inferior',
    'delantero', 'trasero'
}

# Words that are NOT nouns and should be ignored
HARD_STOPWORDS = {
    # Abbrev and direction/position
    'izq', 'der', 'tra', 'ant', 'post', 'sup', 'inf', 'rh', 'lh',
    # Determiners / prepositions
    'de', 'del', 'la', 'el', 'los', 'las', 'un', 'una', 'por', 'para', 'con', 'sin',
    # Generic UI/ops words occasionally present
    'kit', 'set', 'par', 'nuevo', 'nueva', 'usado', 'base', 'tot'
}

# Confident irregulars and heuristics
FEM_EXCEPTIONS = {'mano', 'radio', 'foto'}
MASC_EXCEPTIONS = {'dia', 'mapa', 'planeta', 'problema', 'sistema'}

# Additional suffix rules (very conservative)
FEM_SUFFIXES = ('cion', 'sion', 'xion', 'dad', 'tud', 'umbre', 'sis')
MASC_SUFFIXES = ('aje', 'or')


def is_plausible_noun(tok: str, abbrev_keys: set[str]) -> bool:
    t = tok.lower()
    if not t or len(t) <= 2:
        return False
    # reject numbers and mixed alphanumerics (part numbers)
    if t.isdigit() or re.search(r"[a-z][0-9]|[0-9][a-z]", t):
        return False
    # reject if an abbreviation key or known non-noun
    if t in abbrev_keys or t in HARD_STOPWORDS:
        return False
    # reject if looks like pure adjective (directional etc.)
    if t in DIR_ADJ:
        return False
    # letters only is fine; otherwise, discard
    if not re.fullmatch(r"[a-záéíóúñ]+", t):
        return False
    return True


def guess_gender(word: str) -> str | None:
    w = word.lower()
    g = _gender_from_lists(w)
    if g == 'masculine':
        return 'm'
    if g == 'feminine':
        return 'f'
    # irregulars
    if w in FEM_EXCEPTIONS:
        return 'f'
    if w in MASC_EXCEPTIONS:
        return 'm'
    # conservative morphology
    if w.endswith('a') and not w.endswith(('ma', 'pa', 'ta')):
        return 'f'
    if w.endswith('o'):
        return 'm'
    if w.endswith(FEM_SUFFIXES):
        return 'f'
    if w.endswith(MASC_SUFFIXES):
        return 'm'
    return None


def load_rules_context(xlsx_path: Path) -> tuple[set[str], set[str]]:
    """Return (abbrev_keys, canonical_nouns).
    - abbrev_keys: tokens in Abbreviations sheet (keys), used for filtering.
    - canonical_nouns: first-column tokens from Equivalencias sheet (single-token only),
      used to boost noun confidence.
    """
    import openpyxl
    abbrev_keys: set[str] = set()
    canonical: set[str] = set()
    if not xlsx_path.exists():
        return abbrev_keys, canonical
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'Abbreviations' in wb.sheetnames:
        sh = wb['Abbreviations']
        for row in sh.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip().lower()
            if key:
                abbrev_keys.add(key)
    if 'Equivalencias' in wb.sheetnames:
        sh = wb['Equivalencias']
        for row in sh.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            first = str(row[0]).strip().lower()
            if first and re.fullmatch(r"[a-záéíóúñ]+", first):
                canonical.add(first)
    return abbrev_keys, canonical


def extract_candidates(limit: int | None, abbrev_keys: set[str], canonical: set[str]) -> Counter:
    if not DB.exists():
        raise SystemExit(f"DB not found: {DB}")
    con = sqlite3.connect(DB)
    cur = con.cursor()
    q = "SELECT normalized_descripcion FROM processed_consolidado WHERE normalized_descripcion IS NOT NULL"
    if limit is not None:
        q += f" LIMIT {int(limit)}"

    right_after = Counter()
    after_article = Counter()

    ARTICLES = {'el', 'la', 'los', 'las', 'un', 'una'}

    for (text,) in cur.execute(q):
        if not text:
            continue
        s = str(text).lower().replace('.', ' ').replace('/', ' ')
        toks = TOKEN_RE.findall(s)
        n = len(toks)
        for i, tok in enumerate(toks):
            t = tok.lower()
            if t in DIR_ADJ and i + 1 < n:
                cand = toks[i + 1].lower()
                if is_plausible_noun(cand, abbrev_keys):
                    right_after[cand] += 1
            if t in ARTICLES and i + 1 < n:
                cand = toks[i + 1].lower()
                if is_plausible_noun(cand, abbrev_keys):
                    after_article[cand] += 1

    con.close()

    # Weighted merge; favor article/directional contexts; boost if in canonical list
    out = Counter()
    for w, c in right_after.items():
        out[w] += c * 3
    for w, c in after_article.items():
        out[w] += c * 2
    for w in list(out.keys()):
        if w in canonical:
            out[w] += 5
    return out


def load_existing_sheet(xlsx_path: Path):
    import openpyxl
    if not xlsx_path.exists():
        raise SystemExit(f"Rules file not found: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)
    if 'Noun_Gender' not in wb.sheetnames:
        sh = wb.create_sheet('Noun_Gender')
        sh.cell(1, 1, 'noun')
        sh.cell(1, 2, 'gender')
    else:
        sh = wb['Noun_Gender']
        if (sh.cell(1, 1).value or '').strip().lower() != 'noun':
            sh.cell(1, 1, 'noun')
        if (sh.cell(1, 2).value or '').strip().lower() != 'gender':
            sh.cell(1, 2, 'gender')
    existing: dict[str, str] = {}
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        noun = str(row[0]).strip().lower()
        gender = (str(row[1]).strip().lower() if row[1] is not None else '')
        existing[noun] = gender
    return wb, sh, existing


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--top', type=int, default=None, help='approx max nouns to append (and scale of scan)')
    ap.add_argument('--all', action='store_true', help='scan all rows and consider all candidates')
    args = ap.parse_args()

    abbrev_keys, canonical = load_rules_context(XLSX)

    if args.all:
        scan_limit = None
        top_n = None
    else:
        top_n = args.top or 8000
        scan_limit = (args.top or 8000) * 5

    candidates = extract_candidates(scan_limit, abbrev_keys, canonical)

    wb, sh, existing = load_existing_sheet(XLSX)

    added = 0
    # Order by score desc, then alphabetically
    for noun, _score in (candidates.most_common(top_n) if top_n is not None else candidates.most_common()):
        if noun in existing:  # already present -> skip (we append only new info)
            continue
        # Double-check plausibility before writing
        if not is_plausible_noun(noun, abbrev_keys):
            continue
        g = guess_gender(noun)
        if g in ('m', 'f'):
            sh.append([noun, g])
            existing[noun] = g
            added += 1
        # else: skip uncertain nouns to maintain sheet quality

    wb.save(XLSX)
    print(f"Appended {added} new nouns with gender to {XLSX}")


if __name__ == '__main__':
    main()

