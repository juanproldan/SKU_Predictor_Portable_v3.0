#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Unified Consolidado Processor

This script processes Consolidado.json into a single processed_consolidado.db database
that serves both VIN training and SKU prediction needs. It preserves all useful records
while cleaning VINs and normalizing text descriptions.

Key Features:
- Reads Consolidado.json from Source_Files/
- Cleans VINs using validation logic (preserves records with invalid VINs)
- Processes text using normalization rules
- Creates single processed_consolidado.db
- Supports both VIN training and SKU prediction needs
- Uses dynamic path resolution for client deployment

Author: Augment Agent
Date: 2025-07-24
"""

import os
import sys
import json
import sqlite3
import openpyxl
import re
import time
from pathlib import Path
import logging
from datetime import datetime

# Add utils to path for text processing (ensure 'src' is importable)
SRC_DIR = os.path.dirname(__file__)
if SRC_DIR not in sys.path:
    sys.path.insert(0, SRC_DIR)
# Use the single shared text processor; no legacy/text_utils fallbacks
from utils.unified_text import unified_text_preprocessing

# Global variables for text processing maps
user_corrections_map_global = {}
abbreviations_map_global = {}
synonym_expansion_map_global = {}

# --- Configuration ---
def get_base_path():
    """Get the base path for the application, works for both script and executable."""
    if getattr(sys, 'frozen', False):
        # Running as executable - executable is in portable_app root
        return os.path.dirname(sys.executable)
    else:
        # Running as script - script is in src/, go up to portable_app root
        script_dir = os.path.dirname(os.path.abspath(__file__))
        portable_app_root = os.path.dirname(script_dir)  # Go up from src/ to portable_app root
        return portable_app_root

BASE_PATH = get_base_path()
# Canonical client data folder: Source_Files at project root
PROJECT_ROOT = os.path.dirname(BASE_PATH)
DATA_DIR = os.path.join(PROJECT_ROOT, "Source_Files")
SOURCE_FILES_DIR = DATA_DIR  # for logging
LOGS_DIR = os.path.join(BASE_PATH, "logs")

# File paths - everything in Source_Files for client structure
CONSOLIDADO_PATH = os.path.join(DATA_DIR, "Consolidado.json")
TEXT_PROCESSING_PATH = os.path.join(DATA_DIR, "Text_Processing_Rules.xlsx")
WMI_CSV_PATH = os.path.join(DATA_DIR, "WMI.csv")
OUTPUT_DB_PATH = os.path.join(DATA_DIR, "processed_consolidado.db")
LOG_PATH = os.path.join(LOGS_DIR, f"consolidado_processing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

# --- Logging Setup ---
def setup_logging(verbose: bool = False):
    """Setup performance-optimized logging configuration."""
    try:
        # Try to import from utils module (works when running as script from src/)
        import sys
        script_dir = os.path.dirname(os.path.abspath(__file__))
        if script_dir not in sys.path:
            sys.path.insert(0, script_dir)

        from utils.logging_config import get_logger, create_processing_config, log_operation_start
        config = create_processing_config(verbose=verbose)
        return get_logger("consolidado_processor", config)
    except ImportError:
        # Fallback to basic logging if logging_config not available
        os.makedirs(LOGS_DIR, exist_ok=True)
        logging.basicConfig(
            level=logging.WARNING,  # Reduced from INFO to WARNING
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(LOG_PATH, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        return logging.getLogger(__name__)

# --- VIN Cleaning Functions ---

def canonicalize_vin_chars(vin_str: str) -> str:
    """Apply VIN character corrections: I->1, O->0, Q->0, then uppercase."""
    v = str(vin_str).strip().upper()
    return v.replace('I', '1').replace('O', '0').replace('Q', '0')

def validate_vin_format(vin_str):
    """
    Basic VIN format validation.
    Returns True if VIN has correct format (17 alphanumeric characters).
    """
    if not vin_str or len(vin_str) != 17:
        return False

    # After canonicalization, VIN should only contain A-H, J-N, P, R-Z, 0-9
    import re
    if not re.match(r'^[A-HJ-NPR-Z0-9]{17}$', vin_str):
        return False

    # Must be alphanumeric
    if not vin_str.isalnum():
        return False

    return True

def validate_vin_check_digit(vin_str):
    """
    Validate VIN check digit (position 9) using the standard algorithm.
    This is optional validation - some VINs may have incorrect check digits
    but still be valid for maker/model/series extraction.
    """
    # VIN character values for check digit calculation
    char_values = {
        'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7, 'H': 8,
        'J': 1, 'K': 2, 'L': 3, 'M': 4, 'N': 5, 'P': 7, 'R': 9, 'S': 2,
        'T': 3, 'U': 4, 'V': 5, 'W': 6, 'X': 7, 'Y': 8, 'Z': 9,
        '0': 0, '1': 1, '2': 2, '3': 3, '4': 4, '5': 5, '6': 6, '7': 7, '8': 8, '9': 9
    }

    # Position weights for check digit calculation
    weights = [8, 7, 6, 5, 4, 3, 2, 10, 0, 9, 8, 7, 6, 5, 4, 3, 2]

    try:
        total = sum(char_values[char] * weight for char, weight in zip(vin_str, weights))
        calculated_check = total % 11
        expected_check = 'X' if calculated_check == 10 else str(calculated_check)
        return vin_str[8] == expected_check
    except (KeyError, IndexError):
        return False

def clean_vin_for_training(vin):
    """
    Clean and validate VIN for training purposes.
    Returns cleaned VIN if valid, None if invalid.

    Rules:
    - Uppercase and replace I→1, O/Q→0
    - Reject if not length 17 or contains invalid charset
    - Reject if any character repeats >4 times consecutively (noise VINs)
    """
    if not vin:
        return None

    # Convert to string and clean, with canonicalization
    vin_str = canonicalize_vin_chars(vin)

    # Basic format validation
    if not validate_vin_format(vin_str):
        return None

    # Reject VINs with any character repeated >4 times consecutively
    import re
    if re.search(r'(.)\1{4,}', vin_str):
        return None

    # Optional: Check digit validation (commented out as it may be too strict)
    # if not validate_vin_check_digit(vin_str):
    #     return None

    return vin_str

# --- Database Setup ---
def setup_database(db_path):
    """
    Sets up the SQLite database and the processed_consolidado table.
    IMPORTANT: This function completely replaces any existing database to ensure fresh data.
    """
    logger = logging.getLogger(__name__)
    logger.info(f"Setting up database at: {db_path}")

    try:
        # Ensure Source_Files directory exists (unified structure)
        os.makedirs(os.path.dirname(db_path), exist_ok=True)

        # Remove existing database file if it exists to ensure fresh data
        if os.path.exists(db_path):
            logger.info(f"Removing existing database to ensure fresh data: {db_path}")
            try:
                os.remove(db_path)
            except PermissionError as e:
                logger.error(f"Cannot remove database file (in use by another process): {e}")
                logger.info("Attempting to create new database with temporary name...")
                # Create a backup name and try again
                import time
                backup_name = f"{db_path}.backup_{int(time.time())}"
                try:
                    os.rename(db_path, backup_name)
                    logger.info(f"Moved existing database to: {backup_name}")
                except Exception as e2:
                    logger.error(f"Cannot move database file either: {e2}")
                    raise Exception(f"Database file is locked by another process. Please close all applications using the database and try again.")

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Create the unified table - YEAR RANGE OPTIMIZATION SCHEMA
        # Note: Using CREATE TABLE (not IF NOT EXISTS) since we deleted the file above
        cursor.execute('''
        CREATE TABLE processed_consolidado (
            vin_number TEXT,                    -- Cleaned VINs for VIN training (may be NULL)
            maker TEXT,                         -- For both VIN & SKU training
            model INTEGER,                      -- For both VIN & SKU training
            series TEXT,                        -- For both VIN & SKU training
            descripcion TEXT,                   -- Original description from consolidado.json (may be NULL)
            normalized_descripcion TEXT,        -- Normalized description for SKU training (may be NULL)
            referencia TEXT,                    -- For SKU training (may be NULL)
            UNIQUE(vin_number, descripcion, referencia) -- Prevent duplicates
        )
        ''')

        # Drop existing year range tables if they exist (for clean rebuild)
        cursor.execute('DROP TABLE IF EXISTS sku_year_ranges')

        # Create aggregated year range tables for improved frequency counting
        cursor.execute('''
        CREATE TABLE sku_year_ranges (
            maker TEXT,
            series TEXT,
            descripcion TEXT,
            normalized_descripcion TEXT,
            referencia TEXT,
            start_year INTEGER,
            end_year INTEGER,
            frequency INTEGER,
            global_sku_frequency INTEGER,  -- How many times this SKU appears in entire consolidado
            PRIMARY KEY (maker, series, descripcion, referencia)
        )
        ''')



        # Create indexes for better query performance
        cursor.execute('CREATE INDEX idx_vin_training ON processed_consolidado (vin_number, maker, model, series)')
        cursor.execute('CREATE INDEX idx_referencia_training ON processed_consolidado (maker, model, series, referencia)')
        cursor.execute('CREATE INDEX idx_exact_match ON processed_consolidado (maker, model, series, normalized_descripcion)')
        cursor.execute('CREATE INDEX idx_description_search ON processed_consolidado (normalized_descripcion)')

        # Create indexes for year range tables
        cursor.execute('CREATE INDEX idx_sku_year_range_lookup ON sku_year_ranges (maker, series, start_year, end_year)')
        cursor.execute('CREATE INDEX idx_sku_frequency ON sku_year_ranges (frequency DESC)')

        conn.commit()
        logger.info("Fresh database and 'processed_consolidado' table created successfully.")
        return conn

    except Exception as e:
        logger.error(f"Error setting up database: {e}")
        if 'conn' in locals():
            conn.close()
        return None

# --- Text Processing ---
def load_equivalencias_map(text_processing_path):
    """
    Load equivalencias mapping from Text_Processing_Rules.xlsx.
    Returns dictionary mapping original terms to normalized terms.
    """
    logger = logging.getLogger(__name__)

    if not os.path.exists(text_processing_path):
        logger.warning(f"Text processing rules file not found at {text_processing_path}")
        return {}

    try:
        wb = openpyxl.load_workbook(text_processing_path, data_only=True)
        if 'Equivalencias' not in wb.sheetnames:
            raise FileNotFoundError("Missing 'Equivalencias' sheet in Text_Processing_Rules.xlsx")
        sh = wb['Equivalencias']
        headers = [c.value for c in next(sh.iter_rows(min_row=1, max_row=1))]
        col_idx = {h: i for i, h in enumerate(headers)}
        equivalencias_map = {}
        for row in sh.iter_rows(min_row=2, values_only=True):
            original = str(row[col_idx.get('Original', 0)] or '').strip().upper()
            normalized = str(row[col_idx.get('Normalized', 1)] or '').strip().upper()
            if original and normalized:
                equivalencias_map[original] = normalized
        logger.info(f"Loaded {len(equivalencias_map)} equivalencias mappings")
        return equivalencias_map

    except Exception as e:
        logger.error(f"Error loading equivalencias map: {e}")
        raise

def load_user_corrections_map(text_processing_path):
    """
    Load user corrections mapping from Text_Processing_Rules.xlsx User_Corrections tab.
    Returns dictionary mapping original text to corrected text.
    """
    logger = logging.getLogger(__name__)

    if not os.path.exists(text_processing_path):
        logger.warning(f"Text processing rules file not found at {text_processing_path}")
        return {}

    try:
        wb = openpyxl.load_workbook(text_processing_path, data_only=True)
        sh = wb['User_Corrections']
        headers = [c.value for c in next(sh.iter_rows(min_row=1, max_row=1))]
        col_idx = {h: i for i, h in enumerate(headers)}
        corrections_map = {}
        for row in sh.iter_rows(min_row=2, values_only=True):
            orig = row[col_idx.get('Original_Text', 0)]
            corr = row[col_idx.get('Corrected_Text', 1)]
            if orig is not None and corr is not None:
                original = str(orig).strip()
                corrected = str(corr).strip()
                if original and corrected:
                    corrections_map[original] = corrected
        logger.info(f"Loaded {len(corrections_map)} user corrections")
        return corrections_map

    except Exception as e:
        logger.error(f"Error loading user corrections map: {e}")
        return {}

def load_abbreviations_map(text_processing_path):
    """
    Load abbreviations mapping from Text_Processing_Rules.xlsx Abbreviations tab.
    Returns dictionary mapping abbreviations to full forms.
    """
    logger = logging.getLogger(__name__)

    if not os.path.exists(text_processing_path):
        logger.warning(f"Text processing rules file not found at {text_processing_path}")
        return {}

    try:
        wb = openpyxl.load_workbook(text_processing_path, data_only=True)
        if 'Abbreviations' not in wb.sheetnames:
            raise FileNotFoundError("Missing 'Abbreviations' sheet in Text_Processing_Rules.xlsx")
        sh = wb['Abbreviations']
        # Read arbitrary-width rows: first cell canonical, rest abbreviations
        abbreviations_map = {}
        for i, row in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
            vals = [str(v).strip().lower() for v in row if v is not None and str(v).strip()]
            if len(vals) >= 2:
                canonical_form, *abbrs = vals
                for abbr in abbrs:
                    abbreviations_map[abbr] = canonical_form
        logger.info(f"Loaded {len(abbreviations_map)} abbreviations")
        return abbreviations_map

    except Exception as e:
        logger.error(f"Error loading abbreviations map: {e}")
        raise

def load_series_normalization_map(text_processing_path):
    """
    Load series normalization mapping from Text_Processing_Rules.xlsx Series tab.
    Returns dictionary mapping (maker, series) to normalized_series.
    """
    logger = logging.getLogger(__name__)

    if not os.path.exists(text_processing_path):
        logger.warning(f"Text processing rules file not found at {text_processing_path}")
        return {}

    try:
        wb = openpyxl.load_workbook(text_processing_path, data_only=True)
        sh = wb['Series']
        series_map = {}
        # Iterate rows; collect non-empty cells
        for row in sh.iter_rows(min_row=2, values_only=True):
            series_variations = [str(v).strip() for v in row if v is not None and str(v).strip()]

            if len(series_variations) < 2:
                continue  # Need at least 2 variations to create mappings

            # First variation is the canonical/normalized form
            normalized_series = series_variations[0]

            # Extract maker from the normalized form if it contains maker info
            # Example: "MAZDA/CX-30 (DM)/BASICO" -> maker="MAZDA", series="CX-30"
            if '/' in normalized_series:
                parts = normalized_series.split('/')
                if len(parts) >= 2:
                    maker = parts[0].strip()
                    series_part = parts[1].strip()
                    # Remove parenthetical info: "CX-30 (DM)" -> "CX-30"
                    if '(' in series_part:
                        series_part = series_part.split('(')[0].strip()
                    normalized_series = series_part
                else:
                    # Fallback: use the whole string and assume no specific maker
                    maker = None
            else:
                # Simple series without maker info
                maker = None

            # Create mappings for all variations
            for variation in series_variations:
                if variation != normalized_series:  # Don't map to itself
                    # Clean the variation (remove maker prefix if present)
                    clean_variation = variation
                    if '/' in variation:
                        parts = variation.split('/')
                        if len(parts) >= 2:
                            clean_variation = parts[1].strip()
                            if '(' in clean_variation:
                                clean_variation = clean_variation.split('(')[0].strip()

                    # Create mapping key
                    if maker:
                        # Maker-specific mapping
                        key = (maker.upper(), clean_variation.upper())
                        series_map[key] = normalized_series
                        logger.debug(f"Series mapping: {maker}/{clean_variation} → {normalized_series}")
                    else:
                        # Generic mapping (applies to all makers)
                        key = ("*", clean_variation.upper())
                        series_map[key] = normalized_series
                        logger.debug(f"Series mapping: */{clean_variation} → {normalized_series}")

        logger.info(f"Loaded {len(series_map)} series normalization mappings")
        return series_map

    except Exception as e:
        logger.error(f"Could not load series normalization (Series tab may not exist): {e}")
        raise

def normalize_series_preprocessing(maker, series, series_map):
    """
    Normalize series using the series normalization mapping during preprocessing.

    Args:
        maker: Vehicle maker (e.g., "Mazda", "Ford")
        series: Original series (e.g., "CX30", "CX 30")
        series_map: Series normalization mapping dictionary

    Returns:
        Normalized series (e.g., "CX-30") or original if no mapping found
    """
    if not series or not series_map:
        return series

    # Clean inputs
    maker_clean = maker.upper().strip() if maker else "*"
    series_clean = series.upper().strip()

    # Try maker-specific mapping first
    maker_key = (maker_clean, series_clean)
    if maker_key in series_map:
        normalized = series_map[maker_key]
        logging.getLogger(__name__).debug(f"Series normalized: {maker}/{series} → {normalized} (maker-specific)")
        return normalized

    # Try generic mapping (applies to all makers)
    generic_key = ("*", series_clean)
    if generic_key in series_map:
        normalized = series_map[generic_key]
        logging.getLogger(__name__).debug(f"Series normalized: {maker}/{series} → {normalized} (generic)")
        return normalized

    # No mapping found, return original
    return series

# --- Unified Text Processing Functions ---
def apply_user_corrections(text: str) -> str:
    """
    Apply user corrections from the User_Corrections tab.
    This has the HIGHEST priority in text processing.
    """
    if not text or not user_corrections_map_global:
        return text

    # Check for exact phrase match first (highest priority)
    if text in user_corrections_map_global:
        corrected = user_corrections_map_global[text]
        # Removed verbose logging for performance
        return corrected

    return text

def apply_abbreviations(text: str) -> str:
    """
    Apply abbreviations expansion from the Abbreviations tab with context-aware logic.
    """
    if not text or not abbreviations_map_global:
        return text

    # Split on both spaces and dots to handle cases like "BOC.INF.PUER.DEL.I."
    import re
    words = re.split(r'[\s.]+', text)
    words = [w for w in words if w]  # Remove empty strings
    expanded_words = []

    for i, word in enumerate(words):
        # Clean word for lookup (remove punctuation, convert to lowercase)
        clean_word = word.lower().strip('.,;:!?')

        # Context-aware abbreviation expansion
        should_expand = True

        # Special handling for common prepositions that shouldn't be expanded
        if clean_word in ['de', 'd']:
            # Don't expand "DE" if it's clearly a preposition
            # Look at surrounding context
            prev_word = words[i-1].lower() if i > 0 else ""
            next_word = words[i+1].lower() if i < len(words)-1 else ""

            # "DE" is likely a preposition in these contexts
            preposition_contexts = [
                'absorbedor', 'amortiguador', 'soporte', 'base', 'tapa', 'cubierta',
                'protector', 'guardapolvo', 'sello', 'junta', 'empaque', 'filtro',
                'bomba', 'motor', 'sensor', 'valvula', 'tubo', 'manguera',
                'impactos', 'choque', 'golpe', 'suspension', 'direccion'
            ]

            if (prev_word in preposition_contexts or
                next_word in preposition_contexts or
                next_word.startswith('impact')):
                should_expand = False
                # Removed verbose logging for performance

        if should_expand and clean_word in abbreviations_map_global:
            expanded = abbreviations_map_global[clean_word]
            expanded_words.append(expanded)
            # Removed verbose logging for performance
        else:
            expanded_words.append(word)

    return ' '.join(expanded_words)

def expand_synonyms(text: str) -> str:
    """
    Global synonym expansion function that replaces industry-specific synonyms
    with their equivalence group representatives.

    Note: This function does NOT normalize text - that's done later in the pipeline.
    """
    if not text or not synonym_expansion_map_global:
        return text

    # Split text into words for synonym checking (no normalization here)
    words = text.split()
    expanded_words = []

    for word in words:
        # Check if this word has an industry-specific synonym in Equivalencias (CASE-INSENSITIVE)
        word_lower = word.lower()
        if word_lower in synonym_expansion_map_global:
            group_id = synonym_expansion_map_global[word_lower]
            # Use the group_id as a consistent representation
            group_representative = f"GROUP_{group_id}"
            expanded_words.append(group_representative)
            logging.getLogger(__name__).debug(f"Industry synonym: '{word}' -> '{group_representative}' (Group ID: {group_id})")
        else:
            expanded_words.append(word)

    return ' '.join(expanded_words)

# Replace local pipeline with import of the canonical processor to guarantee parity
from utils.unified_text import unified_text_preprocessing  # noqa: F401

# --- Year Range Aggregation Functions ---
def detect_year_ranges(years):
    """
    Detect year ranges from a list of years, allowing for 1-2 year gaps.

    Args:
        years: List of integer years

    Returns:
        tuple: (start_year, end_year) representing the full range

    Example:
        [2012, 2013, 2015, 2016, 2018] -> (2012, 2018)
        [2020, 2021, 2022] -> (2020, 2022)
    """
    if not years:
        return None, None

    years = sorted(set(years))  # Remove duplicates and sort

    if len(years) == 1:
        return years[0], years[0]

    # For automotive parts, create one continuous range from min to max
    # This handles the reality that parts work in year ranges, not individual years
    # Gaps likely occur due to missing bid data, not actual incompatibility
    start_year = min(years)
    end_year = max(years)

    return start_year, end_year


def aggregate_sku_year_ranges(conn):
    """
    Aggregate SKU data into year ranges for improved frequency counting.

    This function:
    1. Groups SKU records by (maker, series, descripcion, referencia)
    2. Detects year ranges for each group
    3. Calculates total frequency across the year range
    4. Inserts aggregated data into sku_year_ranges table
    """
    logger = logging.getLogger(__name__)
    logger.info("🔄 Aggregating SKU data into year ranges...")

    cursor = conn.cursor()

    # First, calculate global SKU frequencies (how many times each SKU appears in entire consolidado)
    cursor.execute("""
        SELECT referencia, COUNT(*) as global_frequency
        FROM processed_consolidado
        WHERE referencia IS NOT NULL
        AND referencia != ''
        AND referencia != 'None'
        AND referencia != 'UNKNOWN'
        GROUP BY referencia
    """)

    global_sku_frequencies = dict(cursor.fetchall())
    logger.info(f"📊 Calculated global frequencies for {len(global_sku_frequencies):,} unique SKUs")

    # Get all SKU combinations with their years and frequencies
    # Enforce valid model year bounds in aggregation
    from datetime import datetime as _dt
    _min_year = 1990
    _max_year = _dt.now().year + 2
    cursor.execute("""
        SELECT maker, series, descripcion, normalized_descripcion, referencia, model, COUNT(*) as frequency
        FROM processed_consolidado
        WHERE referencia IS NOT NULL
        AND referencia != ''
        AND referencia != 'None'
        AND referencia != 'UNKNOWN'
        AND maker IS NOT NULL
        AND series IS NOT NULL
        AND model IS NOT NULL
        AND model BETWEEN ? AND ?
        GROUP BY maker, series, descripcion, normalized_descripcion, referencia, model
        ORDER BY maker, series, referencia, model
    """, (_min_year, _max_year))

    raw_data = cursor.fetchall()
    logger.info(f"📊 Processing {len(raw_data):,} individual year records for SKU aggregation")

    # Group by (maker, series, descripcion, referencia) and collect years
    sku_groups = {}
    for row in raw_data:
        maker, series, descripcion, normalized_descripcion, referencia, year, frequency = row

        key = (maker, series, descripcion, normalized_descripcion, referencia)
        if key not in sku_groups:
            sku_groups[key] = {'years': [], 'total_frequency': 0}

        # Add year multiple times based on frequency (to preserve frequency weighting)
        sku_groups[key]['years'].extend([year] * frequency)
        sku_groups[key]['total_frequency'] += frequency

    logger.info(f"📈 Grouped into {len(sku_groups):,} unique SKU combinations")

    # Process each group and insert year ranges
    aggregated_count = 0
    for (maker, series, descripcion, normalized_descripcion, referencia), data in sku_groups.items():
        years = data['years']
        total_frequency = data['total_frequency']

        start_year, end_year = detect_year_ranges(years)

        if start_year is not None and end_year is not None:
            try:
                # Get global frequency for this SKU
                global_freq = global_sku_frequencies.get(referencia, 0)

                cursor.execute("""
                    INSERT OR REPLACE INTO sku_year_ranges
                    (maker, series, descripcion, normalized_descripcion, referencia, start_year, end_year, frequency, global_sku_frequency)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (maker, series, descripcion, normalized_descripcion, referencia, start_year, end_year, total_frequency, global_freq))
                aggregated_count += 1
            except Exception as e:
                logger.warning(f"Error inserting SKU year range: {e}")

    conn.commit()
    logger.info(f"✅ Created {aggregated_count:,} SKU year range records")

    return aggregated_count


# --- VIN prefix frequency table (for VIN analysis and QA) ---
def build_vin_prefix_frequencies(conn: sqlite3.Connection) -> int:
    """Create table with VIN prefix (first 11 chars masked) frequencies per maker/model/series.
    Uses strict VIN validation: length==17, allowed charset (no I/O/Q), and no character
    repeated more than 4 times consecutively. Counts DISTINCT VINs to avoid per-item
    multiplication. Returns number of rows inserted.
    """
    import re
    cur = conn.cursor()

    def _is_valid_vin(vin: str) -> int:
        if not vin:
            return 0
        try:
            v = str(vin).strip().upper()
        except Exception:
            return 0
        if len(v) != 17:
            return 0
        # Allowed characters; exclude I, O, Q
        if not re.match(r'^[A-HJ-NPR-Z0-9]{17}$', v):
            return 0
        # Reject VINs with any character repeated >4 times consecutively
        if re.search(r'(.)\1{4,}', v):
            return 0
        return 1

    # Register VIN validator as SQLite UDF
    conn.create_function("is_valid_vin", 1, _is_valid_vin)

    # Rebuild table
    cur.execute('DROP TABLE IF EXISTS vin_prefix_frequencies')
    cur.execute(
        '''
        CREATE TABLE vin_prefix_frequencies (
            vin_mask TEXT,
            maker TEXT,
            model INTEGER,
            series TEXT,
            frequency INTEGER,
            PRIMARY KEY (vin_mask, maker, model, series)
        )
        '''
    )

    cur.execute(
        '''
        INSERT OR REPLACE INTO vin_prefix_frequencies (vin_mask, maker, model, series, frequency)
        SELECT SUBSTR(UPPER(vin_number), 1, 11) || 'XXXXXX' AS vin_mask,
               maker, model, series,
               COUNT(DISTINCT UPPER(vin_number)) AS frequency
        FROM processed_consolidado
        WHERE vin_number IS NOT NULL
          AND maker IS NOT NULL AND model IS NOT NULL AND series IS NOT NULL
          AND is_valid_vin(vin_number)
        GROUP BY vin_mask, maker, model, series
        '''
    )

    # Index for lookup performance
    cur.execute('CREATE INDEX idx_vin_mask_lookup ON vin_prefix_frequencies (vin_mask, maker, model, series)')
    conn.commit()

    cur.execute('SELECT COUNT(*) FROM vin_prefix_frequencies')
    return cur.fetchone()[0]




# --- Main Processing Logic ---
def process_consolidado_record(record, series_map=None):
    """
    Process a single record from Consolidado.json.
    Returns processed record data or None if record should be skipped.
    Normalization rules (centralized here for consistency):
    - VIN: uppercase with I→1, O/Q→0 (via clean_vin_for_training)
    - maker, series, descripcion, normalized_descripcion: lowercase
    """
    # Extract required fields only (removed unused columns)
    vin = record.get('vin_number')
    make = record.get('maker')
    year = record.get('model')
    series = record.get('series')
    description = record.get('item_original_descripcion')
    referencia = record.get('item_referencia')

    # Clean VIN if present (but don't discard record if invalid)
    cleaned_vin = clean_vin_for_training(vin) if vin else None

    # Clean and validate other fields
    make = str(make).strip().lower() if make else None
    year = int(year) if year and str(year).isdigit() else None
    series = str(series).strip().lower() if series else None
    description = str(description).strip().lower() if description else None
    referencia = str(referencia).strip() if referencia and str(referencia).strip() else None

    # Apply series normalization during preprocessing (hybrid approach - Phase 1)
    if series and series_map:
        normalized_series = normalize_series_preprocessing(make, series, series_map)
        if normalized_series != series:
            logging.getLogger(__name__).info(f"Series preprocessed: {make}/{series} → {normalized_series}")
            series = normalized_series.lower() if normalized_series else series

    # Determine record usefulness
    good_for_vin_training = (cleaned_vin and make and year and series)
    good_for_sku_training = (make and year and series and description and referencia)

    if not (good_for_vin_training or good_for_sku_training):
        return None  # Skip - not useful for either training purpose

    # Process description if present
    descripcion = description  # lowercase original description
    normalized_descripcion = None

    if description:
        # Any failure to normalize should be surfaced and stop the pipeline
        normalized_descripcion = unified_text_preprocessing(description)
        if normalized_descripcion:
            normalized_descripcion = str(normalized_descripcion).lower()

    return {
        'vin_number': cleaned_vin,
        'maker': make,
        'model': year,
        'series': series,
        'descripcion': descripcion,
        'normalized_descripcion': normalized_descripcion,
        'referencia': referencia
    }

def process_consolidado_to_db(conn, consolidado_path, series_map=None):
    """
    Reads Consolidado.json, processes records, and inserts into the database.
    """
    logger = logging.getLogger(__name__)
    logger.info(f"Processing consolidado data from: {consolidado_path}")

    if not os.path.exists(consolidado_path):
        logger.error(f"Consolidado file not found at {consolidado_path}")
        return False

    # Statistics
    stats = {
        'total_records': 0,
        'total_items': 0,
        'inserted_records': 0,
        'skipped_insufficient_data': 0,
        'skipped_duplicates': 0,
        'vin_training_records': 0,
        'sku_training_records': 0,
        'both_training_records': 0
    }

    try:
        logger.info("Loading Consolidado.json...")
        with open(consolidado_path, 'r', encoding='utf-8') as f:
            all_records = json.load(f)

        stats['total_records'] = len(all_records)
        logger.info(f"Loaded {stats['total_records']} records from Consolidado.json")

        cursor = conn.cursor()

        # Start processing timer
        processing_start_time = time.time()

        # Process each record with progress bar
        progress_bar = None
        try:
            # Try to create progress bar
            import sys
            script_dir = os.path.dirname(os.path.abspath(__file__))
            if script_dir not in sys.path:
                sys.path.insert(0, script_dir)

            from utils.logging_config import PerformanceLogger
            if hasattr(logger, 'create_progress_bar'):
                progress_bar = logger.create_progress_bar(stats['total_records'], "Processing records")
        except:
            pass

        # Fallback progress reporting
        progress_interval = max(1000, stats['total_records'] // 100)  # Report every 1% or 1000 records, whichever is larger

        for record_idx, record in enumerate(all_records):
            # Update progress bar or fallback to text progress
            if progress_bar:
                progress_bar.update(1)
            elif record_idx % progress_interval == 0 and record_idx > 0:
                elapsed = time.time() - processing_start_time
                rate = record_idx / elapsed if elapsed > 0 else 0
                eta_seconds = (stats['total_records'] - record_idx) / rate if rate > 0 else 0
                eta_minutes = eta_seconds / 60

                progress_pct = (record_idx / stats['total_records']) * 100
                logger.info(f"Progress: {record_idx:,}/{stats['total_records']:,} ({progress_pct:.1f}%) | "
                           f"Rate: {rate:.0f} rec/s | ETA: {eta_minutes:.1f}m")

            # Each record can have multiple items
            items = record.get('items', [])
            stats['total_items'] += len(items)

            # Extract VIN info from record level (simplified - removed unused fields)
            vin_number_raw = record.get('vin_number')
            # Canonicalize VIN characters early (I->1, O/Q->0, uppercase)
            vin_number = canonicalize_vin_chars(vin_number_raw) if vin_number_raw else None

            maker = record.get('maker')  # Field is 'maker'
            model = record.get('model')  # Field is 'model'
            series = record.get('series')  # Field is 'series'

            # If WMI registry is available, log unknown WMIs (first 3 chars)
            try:
                wmi_set = getattr(logger, 'wmi_registry', set())
                if vin_number and len(vin_number) >= 3 and validate_vin_format(vin_number):
                    wmi = vin_number[:3]
                    if wmi_set and wmi not in wmi_set:
                        logger.warning(f"Unknown WMI encountered: {wmi} for VIN prefix {vin_number[:11]}...")
            except Exception:
                pass

            # Process each item in the record
            for item in items:
                # Combine record-level and item-level data (simplified schema)
                combined_record = {
                    'vin_number': vin_number,
                    'maker': maker,
                    'model': model,
                    'series': series,
                    'item_original_descripcion': item.get('descripcion'),  # Field is 'descripcion'
                    'item_referencia': item.get('referencia')  # Field is 'referencia'
                }

                # Process the combined record
                processed_record = process_consolidado_record(combined_record, series_map)

                if processed_record is None:
                    stats['skipped_insufficient_data'] += 1
                    continue

                # Determine training usefulness for statistics
                has_vin_data = (processed_record['vin_number'] and
                               processed_record['maker'] and
                               processed_record['model'] and
                               processed_record['series'])

                has_sku_data = (processed_record['maker'] and
                               processed_record['model'] and
                               processed_record['series'] and
                               processed_record['normalized_descripcion'] and
                               processed_record['referencia'])

                if has_vin_data and has_sku_data:
                    stats['both_training_records'] += 1
                elif has_vin_data:
                    stats['vin_training_records'] += 1
                elif has_sku_data:
                    stats['sku_training_records'] += 1

                # Insert into database
                try:
                    cursor.execute('''
                    INSERT INTO processed_consolidado (
                        vin_number, maker, model, series,
                        descripcion, normalized_descripcion, referencia
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        processed_record['vin_number'],
                        processed_record['maker'],
                        processed_record['model'],
                        processed_record['series'],
                        processed_record['descripcion'],
                        processed_record['normalized_descripcion'],
                        processed_record['referencia']
                    ))
                    stats['inserted_records'] += 1

                except sqlite3.IntegrityError:
                    # Duplicate record
                    stats['skipped_duplicates'] += 1
                except Exception as e:
                    logger.error(f"Error inserting record: {e}")

        # Close progress bar if it was created
        if progress_bar:
            progress_bar.close()

        # Commit all changes
        conn.commit()

        # Calculate processing time and final statistics
        total_processing_time = time.time() - processing_start_time

        # Use performance logger if available
        try:
            from utils.logging_config import log_operation_complete
            processing_stats = {
                'Records Processed': f"{stats['total_records']:,}",
                'Items Processed': f"{stats['total_items']:,}",
                'Records Inserted': f"{stats['inserted_records']:,}",
                'VIN Training Records': f"{stats['vin_training_records']:,}",
                'SKU Training Records': f"{stats['sku_training_records']:,}",
                'Both Training Records': f"{stats['both_training_records']:,}",
                'Skipped (Insufficient)': f"{stats['skipped_insufficient_data']:,}",
                'Skipped (Duplicates)': f"{stats['skipped_duplicates']:,}",
                'Processing Rate': f"{stats['total_records']/total_processing_time:.0f} rec/s"
            }
            log_operation_complete(logger, "Consolidado Processing", total_processing_time, processing_stats)
        except ImportError:
            # Fallback to basic logging
            logger.info("=== Processing Complete ===")
            logger.info(f"Total records processed: {stats['total_records']:,}")
            logger.info(f"Total items processed: {stats['total_items']:,}")
            logger.info(f"Records inserted: {stats['inserted_records']:,}")
            logger.info(f"Processing time: {total_processing_time:.1f}s")
            logger.info(f"Processing rate: {stats['total_records']/total_processing_time:.0f} rec/s")

        return True

    except FileNotFoundError:
        logger.error(f"Consolidado file not found at {consolidado_path}")
        return False
    except json.JSONDecodeError:
        logger.error(f"Could not decode JSON from {consolidado_path}. File might be corrupted.")
        return False
    except MemoryError:
        logger.error(f"MemoryError loading {consolidado_path}. File may be too large.")
        return False
    except Exception as e:
        logger.error(f"Unexpected error occurred: {e}")
        return False

# --- Main Execution ---
def main(verbose: bool = False):
    """Main execution function with performance-optimized logging."""
    import time
    start_time = time.time()

    logger = setup_logging(verbose=verbose)

    # Use performance logger if available
    try:
        from utils.logging_config import log_operation_start, log_operation_complete
        log_operation_start(logger, "Unified Consolidado Processing")
    except ImportError:
        logger.info("=== Starting Unified Consolidado Processing ===")

    if verbose:
        logger.info(f"Base path: {BASE_PATH}")
        logger.info(f"Source files directory: {SOURCE_FILES_DIR}")
        logger.info(f"Consolidado input: {CONSOLIDADO_PATH}")
        logger.info(f"Text processing rules: {TEXT_PROCESSING_PATH}")
        logger.info(f"Database output: {OUTPUT_DB_PATH}")

    # Check required inputs exist
    if not os.path.exists(CONSOLIDADO_PATH):
        logger.error(f"Consolidado.json not found at {CONSOLIDADO_PATH}")
        logger.error("Please ensure Consolidado.json is in the Source_Files directory")
        return False
    if not os.path.exists(TEXT_PROCESSING_PATH):
        logger.error(f"Text_Processing_Rules.xlsx not found at {TEXT_PROCESSING_PATH}")
        return False

    # Load text processing rules (hard-fail on problems)
    logger.info("Loading text processing rules...")
    equivalencias_map = load_equivalencias_map(TEXT_PROCESSING_PATH)

    # Load series normalization rules (hard-fail on problems)
    logger.info("Loading series normalization rules...")
    series_map = load_series_normalization_map(TEXT_PROCESSING_PATH)

    # Load WMI registry if present (for logging unknown WMIs)
    wmi_registry = set()
    if os.path.exists(WMI_CSV_PATH):
        try:
            with open(WMI_CSV_PATH, 'r', encoding='utf-8') as f:
                for line in f:
                    code = line.strip().upper()
                    if code and len(code) >= 3:
                        wmi_registry.add(code[:3])
            logger.info(f"Loaded {len(wmi_registry)} WMI codes from WMI.csv")
        except Exception as e:
            logger.warning(f"Could not read WMI.csv: {e}")

    # Load unified text processing maps (hard-fail on problems)
    logger.info("Loading unified text processing maps...")
    global user_corrections_map_global, abbreviations_map_global, synonym_expansion_map_global

    user_corrections_map_global = load_user_corrections_map(TEXT_PROCESSING_PATH)
    logger.info(f"Loaded {len(user_corrections_map_global)} user corrections")

    abbreviations_map_global = load_abbreviations_map(TEXT_PROCESSING_PATH)
    logger.info(f"Loaded {len(abbreviations_map_global)} abbreviations")

    # Convert equivalencias_map to synonym format for expand_synonyms function
    synonym_expansion_map_global = {}
    for word, group_id in equivalencias_map.items():
        synonym_expansion_map_global[word.lower()] = group_id
    logger.info(f"Loaded {len(synonym_expansion_map_global)} synonyms from equivalencias")

    # spaCy fully removed – no initialization

    # Setup database
    logger.info("Setting up database...")

    # Attach WMI registry to logger for later use
    logger.wmi_registry = wmi_registry
    conn = setup_database(OUTPUT_DB_PATH)
    if not conn:
        logger.error("Failed to setup database. Aborting.")
        return False

    try:
        # Process Consolidado.json
        logger.info("Starting Consolidado.json processing...")
        success = process_consolidado_to_db(conn, CONSOLIDADO_PATH, series_map)

        if success:
            # Aggregate data into year ranges for improved frequency counting (SKU only)
            logger.info("🔄 Starting SKU year range aggregation...")
            sku_ranges_created = aggregate_sku_year_ranges(conn)
            logger.info(f"✅ SKU year range aggregation complete: {sku_ranges_created:,} SKU ranges")

            # Build VIN prefix frequency table for VIN analytics/QA
            logger.info("🔄 Building VIN prefix frequency table...")
            vin_prefix_rows = build_vin_prefix_frequencies(conn)
            logger.info(f"✅ VIN prefix table built: {vin_prefix_rows:,} rows")

            # Final database statistics
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM processed_consolidado")
            total_records = cursor.fetchone()[0]

            cursor.execute("""
                SELECT COUNT(*) FROM processed_consolidado
                WHERE vin_number IS NOT NULL AND maker IS NOT NULL
                AND model IS NOT NULL AND series IS NOT NULL
            """)
            vin_training_ready = cursor.fetchone()[0]

            cursor.execute("""
                SELECT COUNT(*) FROM processed_consolidado
                WHERE maker IS NOT NULL AND model IS NOT NULL
                AND series IS NOT NULL AND descripcion IS NOT NULL
                AND referencia IS NOT NULL
            """)
            sku_training_ready = cursor.fetchone()[0]

            # Year range statistics
            cursor.execute("SELECT COUNT(*) FROM sku_year_ranges")
            sku_year_ranges = cursor.fetchone()[0]



            # Final completion logging
            total_time = time.time() - start_time
            try:
                from utils.logging_config import log_operation_complete
                final_stats = {
                    'Total Records': f"{total_records:,}",
                    'VIN Training Ready': f"{vin_training_ready:,}",
                    'SKU Training Ready': f"{sku_training_ready:,}",
                    'SKU Year Ranges': f"{sku_year_ranges:,}",
                    'VIN Prefix Rows': f"{vin_prefix_rows:,}",
                    'Database Path': OUTPUT_DB_PATH
                }
                log_operation_complete(logger, "Complete Processing Pipeline", total_time, final_stats)
            except ImportError:
                logger.info("=== Final Database Statistics ===")
                logger.info(f"Total records in database: {total_records:,}")
                logger.info(f"Records ready for VIN training: {vin_training_ready:,}")
                logger.info(f"Records ready for SKU training: {sku_training_ready:,}")
                logger.info(f"SKU year ranges created: {sku_year_ranges:,}")
                logger.info(f"VIN prefix frequency rows: {vin_prefix_rows:,}")
                logger.info(f"Total processing time: {total_time:.1f}s")
                logger.info(f"Database created successfully: {OUTPUT_DB_PATH}")

        return success

    finally:
        if conn:
            conn.close()
            if verbose:
                logger.info("Database connection closed.")

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Process Consolidado.json with optimized logging')
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Enable verbose logging (default: optimized for performance)')
    args = parser.parse_args()

    success = main(verbose=args.verbose)
    if not success:
        sys.exit(1)

    print("\n" + "="*60)
    print("CONSOLIDADO PROCESSING COMPLETED SUCCESSFULLY!")
    print("="*60)
    print("Database created: processed_consolidado.db")
    print("Ready for VIN and SKU training")
    print("Check logs for detailed statistics")
    print("="*60)
