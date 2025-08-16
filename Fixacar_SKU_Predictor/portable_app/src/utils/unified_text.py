"""
Unified text, series, and VIN normalization used by both Step 2 (DB build) and Step 5 (GUI).

- unified_text_preprocessing(text): shared description normalization pipeline
- normalize_series(maker, series): apply Series tab rules (map any variant to the first column)
- canonicalize_vin_chars(vin): fix I→1, O/Q→0 and uppercase (never lowercase VINs)

No fallbacks or external NLP deps (no spaCy). All logic is local and rule-based.
"""
from __future__ import annotations
import os
import re
import unicodedata
from typing import Dict, Tuple

# =====================
# Rules loading (cached)
# =====================
_series_map_cache: Dict[str, str] | None = None
_abbrev_map_cache: Dict[str, str] | None = None
_equiv_map_cache: Dict[str, str] | None = None
_usercorr_map_cache: Dict[str, str] | None = None
_noun_gender_map_cache: Dict[str, str] | None = None  # noun (accentless lowercase) -> 'm'|'f'
_phrase_abbrev_map_cache: Dict[str, str] | None = None  # phrase-level expansions


def _rules_path() -> str:
    """Return the single canonical rules file path within the project folder."""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # portable_app/src
    portable_app_dir = os.path.dirname(base_dir)      # portable_app
    project_root = os.path.dirname(portable_app_dir)  # Fixacar_SKU_Predictor
    return os.path.join(project_root, 'Source_Files', 'Text_Processing_Rules.xlsx')


def _load_series_map() -> Dict[str, str]:
    import openpyxl
    path = _rules_path()
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Series' not in wb.sheetnames:
        return {}
    sh = wb['Series']
    series_map: Dict[str, str] = {}
    for row in sh.iter_rows(min_row=2, values_only=True):
        # First non-empty cell is canonical; all other non-empty cells are explicit variants
        raw = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if len(raw) < 2:
            continue
        canonical = raw[0]
        variants = raw[1:]
        for var in variants:
            key = var.upper()
            series_map[key] = canonical
    return series_map


def _load_abbrev_map() -> Dict[str, str]:
    import openpyxl
    path = _rules_path()
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Abbreviations' not in wb.sheetnames:
        return {}
    sh = wb['Abbreviations']

    # Detect layout by headers:
    #  A) canonical-first, then Abbr. 1..N (preferred):  [Word | Abbr. 1 | Abbr. 2 | ...]
    #  B) pair layout: [Abbr | Full]
    h1 = str(sh.cell(1, 1).value or '').strip().lower()
    h2 = str(sh.cell(1, 2).value or '').strip().lower()
    canonical_first = (h1 in {'word', 'canonical', 'full', 'palabra'} or h2.startswith('abbr'))

    m: Dict[str, str] = {}
    for row in sh.iter_rows(min_row=2, values_only=True):
        vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if len(vals) < 2:
            continue
        if canonical_first or len(vals) > 2:
            canonical = _strip_accents(vals[0].lower())
            for ab in vals[1:]:
                key = _strip_accents(str(ab).lower())
                m[key] = canonical
        else:
            # pair form: [abbr, full]
            abbr, full = _strip_accents(vals[0].lower()), _strip_accents(vals[1].lower())
            m[abbr] = full
    return m


def _load_equiv_map() -> Dict[str, str]:
    """Equivalencias: map any synonym/alias to the canonical (first column)."""
    import openpyxl
    path = _rules_path()
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Equivalencias' not in wb.sheetnames:
        return {}
    sh = wb['Equivalencias']
    m: Dict[str, str] = {}
    for row in sh.iter_rows(min_row=2, values_only=True):
        vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if len(vals) >= 2:
            canonical = vals[0].lower()
            for alias in vals[1:]:
                m[str(alias).lower()] = canonical
    return m


def _load_usercorr_map() -> Dict[str, str]:
    """Deprecated: User Corrections are no longer used. Return empty map."""
    return {}


def _load_noun_gender_map() -> Dict[str, str]:
    """Load noun->gender map from new 'Noun_Gender' sheet: columns: noun, gender(m/f).
    Nouns are normalized to lowercase accentless. Gender values to 'm' or 'f'.
    """
    import openpyxl
    path = _rules_path()
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Noun_Gender' not in wb.sheetnames:
        return {}
    sh = wb['Noun_Gender']
    m: Dict[str, str] = {}
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        noun = _strip_accents(str(row[0]).strip().lower())
        g = str(row[1]).strip().lower()
        if g in ('m', 'masculino', 'male'):
            m[noun] = 'm'
        elif g in ('f', 'femenino', 'female'):
            m[noun] = 'f'
    return m


# =====================
# Text normalization
# =====================

def _strip_accents(s: str) -> str:
    s = unicodedata.normalize('NFKD', s)
    return ''.join(ch for ch in s if not unicodedata.combining(ch))


def _basic_tokenize(s: str) -> list[str]:
    return re.findall(r"[a-z0-9]+", s.lower())


_common_gender_pairs = {
    # Intentionally minimal — avoid automatic gender flips that cause errors.
}

# Extra known noun genders to enforce adjective agreement even if sheet lacks the entry
_EXTRA_NOUN_GENDERS = { 'paragolpes': 'm' }




def _gender_normalize_token(tok: str) -> str:
    # Only apply explicit replacements that are clearly adjectival direction/location
    if tok in _common_gender_pairs:
        return _common_gender_pairs[tok]
    return tok


def _apply_user_corrections(text: str) -> str:
    global _usercorr_map_cache
    if _usercorr_map_cache is None:
        _usercorr_map_cache = _load_usercorr_map()
    t = text
    # Apply phrase-level corrections (case-insensitive)
    for src, dst in _usercorr_map_cache.items():
        if not src:
            continue
        t = re.sub(re.escape(src), dst, t, flags=re.IGNORECASE)
    return t


def _expand_abbreviations(text: str) -> str:
    global _abbrev_map_cache
    if _abbrev_map_cache is None:
        _abbrev_map_cache = _load_abbrev_map()
    if not _abbrev_map_cache:
        return text
    # Normalize common separator patterns before tokenizing: treat dot-separated chains as separate tokens
    t = text.replace('.', ' ').replace('/', ' ')
    tokens = _basic_tokenize(t)
    expanded = [(_abbrev_map_cache.get(tok, tok)) for tok in tokens]
    return ' '.join(expanded)


def _load_phrase_abbrev_map() -> Dict[str, str]:
    """Load phrase-level abbreviation mappings from 'Abbreviations_Phrases' sheet.
    Schema: From | To (headers in row 1)
    Case-insensitive; matches are applied before token-level expansion.
    """
    import openpyxl
    path = _rules_path()
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Abbreviations_Phrases' not in wb.sheetnames:
        return {}
    sh = wb['Abbreviations_Phrases']
    m: Dict[str, str] = {}
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        src = _strip_accents(str(row[0]).strip().lower())
        dst = _strip_accents(str(row[1]).strip().lower())
        if src and dst:
            m[src] = dst
    return m


def _apply_equivalencias(text: str) -> str:
    """
    DEPRECATED for normalization: Equivalencias define synonym GROUPS for matching/comparison,
    not canonicalization. This function is kept as a no-op to avoid breaking imports.
    """
    return text


def _apply_adjective_agreement(tokens: list[str], noun_gender: Dict[str, str]) -> list[str]:
    """Adjust direction/location adjectives to match nearest known noun gender.
    We look at a small window around each adjective token and choose 'o'/'a' ending for:
      - izquierdo/izquierda
      - derecho/derecha
      - delantero/delantera
      - trasero/trasera
    We never flip nouns like 'stop'. Only adjust these adjectives.
    """
    ADJ_BASES = {
        'izquierd': ('izquierdo', 'izquierda'),
        'derech': ('derecho', 'derecha'),
        'delanter': ('delantero', 'delantera'),
        'traser': ('trasero', 'trasera'),
    }
    out = tokens[:]
    # Precompute accentless lower tokens for lookup
    base = [_strip_accents(t.lower()) for t in tokens]
    n = len(tokens)

    # Merge external noun genders with extra built-ins
    noun_gender_full = dict(noun_gender)
    noun_gender_full.update(_EXTRA_NOUN_GENDERS)

    for i, bt in enumerate(base):
        # detect adjective forms or their abbreviations expanded earlier
        match = None
        for stem in ADJ_BASES.keys():
            if bt.startswith(stem):
                match = stem
                break
        if not match:
            continue
        # Find nearest noun to the right, else to the left, that exists in noun_gender map
        g = None
        for j in range(i+1, min(i+6, n)):
            w = base[j]
            if w in noun_gender_full:
                g = noun_gender_full[w]
                break
        if g is None:
            for j in range(i-1, max(-1, i-6), -1):
                w = base[j]
                if w in noun_gender_full:
                    g = noun_gender_full[w]
                    break
        if g is None:
            continue
        masc, fem = ADJ_BASES[match]
        out[i] = masc if g == 'm' else fem
    return out


def unified_text_preprocessing(text: str) -> str:
    """Shared description normalization without spaCy or external fallbacks."""
    if text is None or str(text).strip() == '':
        return ''
    # Step 1: user corrections removed — use input as-is (original casing)
    t = str(text)
    # Step 2: lowercase and strip accents for processing
    t = _strip_accents(t.lower())
    # Step 2b: phrase-level abbreviations before tokenization (e.g., "tra d" -> "trasero derecho")
    global _phrase_abbrev_map_cache
    if _phrase_abbrev_map_cache is None:
        _phrase_abbrev_map_cache = _load_phrase_abbrev_map()
    if _phrase_abbrev_map_cache:
        for src, dst in _phrase_abbrev_map_cache.items():
            t = re.sub(rf"(?<![a-z0-9]){re.escape(src)}(?![a-z0-9])", dst, t)
    # Step 3: abbreviations expansion (token-level)
    t = _expand_abbreviations(t)
    # Step 4: RESERVED — Equivalencias are NOT used for normalization.
    #         They are applied only during matching/comparison as synonym groups.
    # Step 5: linguistic normalization (gender heuristics only; NO plural→singular)
    toks = t.split()
    # Do NOT singularize tokens anymore
    # Do not apply generic gender flips
    toks = [_gender_normalize_token(tok) for tok in toks]
    # Step 5b: targeted adjective agreement using Noun_Gender sheet
    global _noun_gender_map_cache
    if _noun_gender_map_cache is None:
        _noun_gender_map_cache = _load_noun_gender_map()
    if _noun_gender_map_cache:
        toks = _apply_adjective_agreement(toks, _noun_gender_map_cache)
    # Step 6: final cleanup
    t = ' '.join(toks)
    t = re.sub(r"\s+", " ", t).strip()
    return t

# =====================
# Series normalization
# =====================

def normalize_series(maker: str | None, series: str | None) -> str | None:
    """Normalize series by replacing only the explicit variant substring with the canonical.
    - Case-insensitive
    - Longest-first order
    - Letter-only variants (e.g., 'CX') will NOT match if followed by optional space/hyphen and digits (prevents 'CX-30' -> 'CX-5-30')
    - All other characters in the string are preserved
    """
    global _series_map_cache
    if series is None or str(series).strip() == '':
        return series
    if _series_map_cache is None:
        _series_map_cache = _load_series_map()

    s_raw = str(series)
    s_up = s_raw.upper()

    def _find_span(key_up: str) -> tuple[int, int] | None:
        # If key is letters only, require token boundary and prevent trailing -/space + digits
        if re.fullmatch(r"[A-Z]+", key_up):
            pattern = r"(^|[^A-Z0-9])(" + re.escape(key_up) + r")(?![ \-]?\d)(?=$|[^A-Z0-9])"
        else:
            # For mixed keys (with digits or hyphens), just require token boundaries
            pattern = r"(^|[^A-Z0-9])(" + re.escape(key_up) + r")(?![A-Z0-9])"
        m = re.search(pattern, s_up)
        if not m:
            return None
        start = m.start(2)
        end = m.end(2)
        return (start, end)

    for key in sorted(_series_map_cache.keys(), key=len, reverse=True):
        key_up = key.upper()
        span = _find_span(key_up)
        if span:
            st, en = span
            canonical = _series_map_cache[key]
            return s_raw[:st] + canonical + s_raw[en:]

    return series

# =====================
# VIN normalization
# =====================

def canonicalize_vin_chars(vin: str | None) -> str | None:
    """Apply VIN character corrections: I→1, O/Q→0, then uppercase. Never lowercase VINs."""
    if vin is None:
        return None
    v = str(vin).strip().upper()
    return v.replace('I', '1').replace('O', '0').replace('Q', '0')

