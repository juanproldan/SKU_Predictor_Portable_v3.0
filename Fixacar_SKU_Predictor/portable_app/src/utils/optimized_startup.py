#!/usr/bin/env python3
"""
Optimized startup components for faster application loading
"""

import os
import json
import pickle
import time
import threading
from pathlib import Path
from typing import Dict, Any, Optional
# No pandas dependency in client-compatible build

# Provide simple accessors used by main_app
_data_loader = None
_model_loader = None
_text_processor = None

def get_data_loader():
    global _data_loader
    if _data_loader is None:
        _data_loader = OptimizedDataLoader()
    return _data_loader


def get_model_loader():
    global _model_loader
    if _model_loader is None:
        _model_loader = OptimizedModelLoader()
    return _model_loader

def get_text_processor(rules: Dict[str, Any] = None):
    global _text_processor
    if _text_processor is None and rules is not None:
        _text_processor = FastTextProcessor(rules)
    return _text_processor

def initialize_optimizations():
    # Initialize globals lazily; nothing to do eagerly
    return True

class OptimizedDataLoader:
    """Optimized data loading with caching and lazy loading"""
    
    def __init__(self, cache_dir: str = None):
        if cache_dir is None:
            # Use utils/performance_improvements/cache as default
            cache_dir = os.path.join(os.path.dirname(__file__), "performance_improvements", "cache")
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self._cache = {}
        self._loading_threads = {}
        
    def get_cache_path(self, source_file: str, cache_type: str = "pickle") -> Path:
        """Get cache file path for a source file"""
        source_path = Path(source_file)
        cache_name = f"{source_path.stem}_{cache_type}.cache"
        return self.cache_dir / cache_name
        
    def is_cache_valid(self, source_file: str, cache_file: Path) -> bool:
        """Check if cache is newer than source file"""
        if not cache_file.exists():
            return False
        
        source_path = Path(source_file)
        if not source_path.exists():
            return False
            
        return cache_file.stat().st_mtime > source_path.stat().st_mtime
        
    def load_excel_optimized(self, excel_file: str, force_refresh: bool = False):
        """Load Excel using openpyxl (pandas-free)."""
        from openpyxl import load_workbook
        wb = load_workbook(excel_file, data_only=True)
        data = {}
        for sh in wb.worksheets:
            rows = list(sh.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h).strip() if h is not None else '' for h in rows[0]]
            records = []
            for r in rows[1:]:
                rec = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
                records.append(rec)
            data[sh.title] = records
        return data

    def load_text_processing_rules_optimized(self, excel_file: str) -> Dict[str, Any]:
        """Load and process text processing rules with optimization (pandas-free)"""
        cache_file = self.get_cache_path(excel_file, "text_rules")

        # Check cache first
        if self.is_cache_valid(excel_file, cache_file):
            try:
                with open(cache_file, 'rb') as f:
                    return pickle.load(f)
            except Exception:
                pass

        # Load and process Excel data (as lists of dicts)
        excel_data = self.load_excel_optimized(excel_file)

        rules = {
            'equivalencias': {},
            'abbreviations': {},
            'user_corrections': {}
        }

        # Process sheets if present
        for rec in excel_data.get('Equivalencias', []):
            orig = rec.get('Original'); eq = rec.get('Equivalencia')
            if orig and eq:
                rules['equivalencias'][str(orig).lower().strip()] = str(eq).lower().strip()
        for rec in excel_data.get('Abbreviations', []):
            ab = rec.get('Abbreviation'); full = rec.get('Full_Form')
            if ab and full:
                rules['abbreviations'][str(ab).lower().strip()] = str(full).lower().strip()
        for rec in excel_data.get('User_Corrections', []):
            orig = rec.get('Original'); corr = rec.get('Corrected')
            if orig and corr:
                rules['user_corrections'][str(orig).lower().strip()] = str(corr).lower().strip()

        # Cache processed rules
        try:
            with open(cache_file, 'wb') as f:
                pickle.dump(rules, f, protocol=pickle.HIGHEST_PROTOCOL)
        except Exception as e:
            print(f"Rules cache write error: {e}")

        return rules

# spaCy fully removed: No class or loading

class OptimizedModelLoader:
    """Optimized model loading with compression and caching"""
    
    def __init__(self, cache_dir: str = None):
        if cache_dir is None:
            # Use utils/performance_improvements/cache as default
            cache_dir = os.path.join(os.path.dirname(__file__), "performance_improvements", "cache")
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self._models = {}
        
    def load_model_optimized(self, model_path: str, model_name: str) -> Any:
        """Load model with optimization"""
        if model_name in self._models:
            return self._models[model_name]
        
        # Check for compressed cache
        cache_file = self.cache_dir / f"{model_name}_compressed.cache"
        
        if cache_file.exists() and self._is_cache_valid(model_path, cache_file):
            try:
                import joblib
                model = joblib.load(cache_file)
                self._models[model_name] = model
                print(f"✅ Loaded {model_name} from compressed cache")
                return model
            except Exception:
                pass
        
        # Load original model
        try:
            import joblib
            model = joblib.load(model_path)
            self._models[model_name] = model
            
            # Create compressed cache
            try:
                joblib.dump(model, cache_file, compress=3)
            except Exception as e:
                print(f"Model cache write error: {e}")
            
            print(f"✅ Loaded {model_name} from original file")
            return model
            
        except Exception as e:
            print(f"❌ Failed to load {model_name}: {e}")
            return None
    
    def _is_cache_valid(self, source_file: str, cache_file: Path) -> bool:
        """Check if cache is valid"""
        source_path = Path(source_file)
        if not source_path.exists():
            return False
        return cache_file.stat().st_mtime > source_path.stat().st_mtime

class FastTextProcessor:
    """Fast text processing without heavy dependencies"""
    
    def __init__(self, rules: Dict[str, Any]):
        self.equivalencias = rules.get('equivalencias', {})
        self.abbreviations = rules.get('abbreviations', {})
        self.user_corrections = rules.get('user_corrections', {})
        
        # Pre-compile common patterns for speed
        self._compile_patterns()
    
    def _compile_patterns(self):
        """Pre-compile regex patterns for common transformations"""
        import re
        
        # Common automotive abbreviations
        self.abbr_patterns = []
        for abbr, full in self.abbreviations.items():
            # Word boundary pattern for accurate replacement
            pattern = re.compile(r'\b' + re.escape(abbr) + r'\b', re.IGNORECASE)
            self.abbr_patterns.append((pattern, full))
    
    def process_fast(self, text: str) -> str:
        """Fast text processing without spaCy"""
        if not text:
            return text
        
        # Normalize
        result = text.lower().strip()
        
        # Apply user corrections first (highest priority)
        for orig, corrected in self.user_corrections.items():
            if orig in result:
                result = result.replace(orig, corrected)
        
        # Apply abbreviations
        for pattern, replacement in self.abbr_patterns:
            result = pattern.sub(replacement, result)
        
        # Apply equivalencias
        for orig, equiv in self.equivalencias.items():
            if orig in result:
                result = result.replace(orig, equiv)
        
        return result.strip()

# Global instances for reuse
_data_loader = None
_model_loader = None
_text_processor = None

def get_data_loader() -> OptimizedDataLoader:
    """Get global data loader instance"""
    global _data_loader
    if _data_loader is None:
        _data_loader = OptimizedDataLoader()
    return _data_loader

def get_model_loader() -> OptimizedModelLoader:
    """Get global model loader instance"""
    global _model_loader
    if _model_loader is None:
        _model_loader = OptimizedModelLoader()
    return _model_loader

def get_text_processor(rules: Dict[str, Any] = None) -> FastTextProcessor:
    """Get global text processor instance"""
    global _text_processor
    if _text_processor is None and rules:
        _text_processor = FastTextProcessor(rules)
    return _text_processor

def initialize_optimizations():
    """Initialize all optimizations"""
    print("🚀 Initializing performance optimizations...")
    
    # spaCy disabled; do not start background loading
    # Initialize other components
    get_data_loader()
    get_model_loader()

    print("✅ Performance optimizations initialized")
